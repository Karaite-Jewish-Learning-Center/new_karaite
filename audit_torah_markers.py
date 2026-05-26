"""Audit Torah Audio Recording Markers.xlsx for bad cells.

A row is flagged if any of these hold:
  - Time format is unparseable.
  - end < start within the same row (negative duration).
  - This row's start is earlier than the previous row's end in the same file
    (sequence regression).
  - end value exceeds the MP3's actual duration (over-runs the file).

Outputs a human-readable report grouped by sheet/file.
"""

from __future__ import annotations

import re
import shutil
import subprocess
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX_PATH = ROOT / "Updated Torah Audio Recordings Markers.xlsx"
SOURCE_AUDIO_DIR = ROOT / "Combined Parashiot by Aliya Folder"
FRAME_RATE = 30.0
TOLERANCE = 1.0  # seconds


def smpte_to_seconds(value):
    if value is None:
        return None, None
    s = str(value).strip()
    if not s:
        return None, None
    parts = s.split(":")
    if len(parts) != 4:
        return None, f"unparseable {s!r}"
    try:
        hh, mm, ss, ff = (int(p) for p in parts)
    except ValueError:
        return None, f"unparseable {s!r}"
    return hh * 3600 + mm * 60 + ss + ff / FRAME_RATE, None


_ALIYAH_RE = re.compile(r"^\s*(\d+)\b.*?\b(\d+)[\s.]*([A-Za-z'_]+)", re.UNICODE)


def normalize_key(name: str) -> str:
    s = re.sub(r"\.mp3$", "", name, flags=re.IGNORECASE)
    m = _ALIYAH_RE.match(s)
    if m:
        return f"{int(m.group(1)):02d}-{int(m.group(2))}"
    s = s.replace("'", "").replace("_", "")
    s = re.sub(r"[^A-Za-z0-9]+", " ", s).strip().lower()
    return re.sub(r"\s+", " ", s)


def disk_index() -> dict[str, str]:
    return {
        normalize_key(p.name): p.name
        for p in SOURCE_AUDIO_DIR.iterdir()
        if p.is_file() and p.suffix.lower() == ".mp3"
    }


def mp3_duration(path: Path) -> float | None:
    try:
        out = subprocess.check_output(
            ["ffprobe", "-v", "error", "-show_entries", "format=duration",
             "-of", "default=nw=1:nk=1", str(path)],
            text=True, stderr=subprocess.DEVNULL,
        )
        return float(out.strip())
    except (subprocess.CalledProcessError, ValueError, FileNotFoundError):
        return None


def audit_sheet(ws, idx_disk):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], {}
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    if "file name" not in header:
        return [], {}

    cols = {name: header.index(name) for name in (
        "file name", "torah portion", "traditional shabbat aliyah",
        "chapter", "verse", "time starts", "time ends",
    ) if name in header}

    findings: list[dict] = []
    durations: dict[str, float | None] = {}
    last_end_by_file: dict[str, float] = {}
    seen_files: set[str] = set()

    for line_no, row in enumerate(rows[1:], start=2):
        filename = row[cols["file name"]]
        chapter = row[cols["chapter"]]
        verse = row[cols["verse"]]
        if not filename or chapter is None or verse is None:
            continue
        filename = str(filename).strip()
        try:
            chapter = int(chapter); verse = int(verse)
        except (TypeError, ValueError):
            continue

        if filename not in seen_files:
            seen_files.add(filename)
            disk_name = idx_disk.get(normalize_key(filename))
            durations[filename] = (
                mp3_duration(SOURCE_AUDIO_DIR / disk_name) if disk_name else None
            )

        start, start_err = smpte_to_seconds(row[cols["time starts"]])
        end, end_err = smpte_to_seconds(row[cols["time ends"]])

        issues = []
        if start_err: issues.append(f"Time Starts {start_err}")
        if end_err: issues.append(f"Time Ends {end_err}")

        if start is not None and end is not None and end < start:
            issues.append(f"end < start ({end:.2f}s vs {start:.2f}s)")

        prev_end = last_end_by_file.get(filename)
        if start is not None and prev_end is not None and start + TOLERANCE < prev_end:
            issues.append(
                f"start regresses below previous end "
                f"(start {start:.2f}s < prev end {prev_end:.2f}s)"
            )

        dur = durations.get(filename)
        if dur is not None:
            if start is not None and start > dur + TOLERANCE:
                issues.append(f"start {start:.2f}s exceeds MP3 duration {dur:.2f}s")
            if end is not None and end > dur + TOLERANCE:
                issues.append(f"end {end:.2f}s exceeds MP3 duration {dur:.2f}s")

        if end is not None:
            last_end_by_file[filename] = max(prev_end or 0, end)

        if issues:
            findings.append({
                "sheet_row": line_no,
                "file": filename,
                "chapter": chapter,
                "verse": verse,
                "raw_start": row[cols["time starts"]],
                "raw_end": row[cols["time ends"]],
                "issues": issues,
            })

    return findings, durations


def main():
    if not XLSX_PATH.exists():
        raise SystemExit(f"Missing xlsx: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    idx_disk = disk_index()

    grand_total = 0
    for sheet in wb.sheetnames:
        findings, durations = audit_sheet(wb[sheet], idx_disk)
        if not findings and not durations:
            print(f"\n=== {sheet} === (no audio markers / nothing to audit)")
            continue

        print(f"\n=== {sheet} ===")
        if not findings:
            print("  no issues found")
            continue

        by_file: dict[str, list[dict]] = {}
        for f in findings:
            by_file.setdefault(f["file"], []).append(f)

        for fname, items in by_file.items():
            dur = durations.get(fname)
            dur_s = f"{dur:.2f}s" if dur is not None else "unknown"
            print(f"\n  {fname}  (mp3 duration: {dur_s})")
            for f in items:
                start = f["raw_start"] if f["raw_start"] is not None else "(empty)"
                end = f["raw_end"] if f["raw_end"] is not None else "(empty)"
                print(
                    f"    row {f['sheet_row']:>4}  "
                    f"Ch {f['chapter']}:{f['verse']}  "
                    f"start={start!s:>15}  end={end!s:>15}"
                )
                for issue in f["issues"]:
                    print(f"        - {issue}")
            grand_total += len(items)

    print(f"\nTotal flagged rows: {grand_total}")


if __name__ == "__main__":
    main()
