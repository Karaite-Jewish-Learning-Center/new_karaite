from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from ...models import (MenuItems,
                       Kedushot)
from django.db import transaction
from pathlib import Path
import os


class Command(BaseCommand):
    help = 'Fill in Kedushot and items database'

    def add_arguments(self, parser):
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Run without saving to database',
        )
        parser.add_argument(
            '--verbose',
            action='store_true',
            help='Show detailed output',
        )

    def handle(self, *args, **options):
        """Import Kedushot and Piyyutim La-Parashiyyot"""
        dry_run = options['dry_run']
        verbose = options['verbose']

        # Check if file exists
        open_file = Path('karaites/management/commands/kedushot.xlsx')
        if not os.path.exists(open_file):
            self.stdout.write(self.style.ERROR(f"Excel file not found at {open_file}"))
            return

        self.stdout.write(self.style.SUCCESS(f"Found Excel file at {open_file}"))

        try:
            wb = load_workbook(open_file)

            # Check if the required sheet exists
            if 'Index Format' not in wb.sheetnames:
                self.stdout.write(self.style.ERROR("Sheet 'Index Format' not found in Excel file"))
                self.stdout.write(f"Available sheets: {wb.sheetnames}")
                return

            index_format_ws = wb['Index Format']
            self.stdout.write(self.style.SUCCESS("Found 'Index Format' sheet"))

            kedushot_order = 1000
            last_kedushot = None
            menu_order = 1000
            weekly_poems = 'Poems for the Weekly Torah Portion'
            weekly_sabbath = 'Poems for the Weekly Sabbath'

            # Count rows for progress reporting
            total_rows = sum(1 for row in index_format_ws.iter_rows(min_row=3) if row[1].value is not None)
            self.stdout.write(f"Found {total_rows} rows to process")

            # Use transaction to ensure all-or-nothing import
            with transaction.atomic():
                # Skip first two rows, read column A,B and C until first empty cell
                row_count = 0
                kedushot_count = 0
                menu_item_count = 0

                for row in index_format_ws.iter_rows(min_row=3, max_col=100, max_row=index_format_ws.max_row):
                    if row[1].value is None:
                        break

                    row_count += 1
                    if verbose:
                        self.stdout.write(
                            f"Processing row {row_count}/{total_rows}: {row[0].value}, {row[1].value}, {row[2].value}")

                    try:
                        if row[0].value in ['Expand/Collapse', '#', 'A', 'B', 'C', 'D', 'E']:
                            belongs_to = weekly_sabbath if row[0].value in [
                                '#', 'A', 'B', 'C', 'D', 'E'] else weekly_poems

                            if dry_run:
                                if verbose:
                                    self.stdout.write(
                                        f"[DRY RUN] Would create Kedushot: {row[1].value} in {belongs_to}")
                                last_kedushot = type('obj', (object,), {
                                    'menu_title_left': row[1].value,
                                    'menu_items': type('obj', (object,), {'add': lambda x: None})
                                })
                            else:
                                last_kedushot, created = Kedushot.objects.get_or_create(
                                    menu_title_left=row[1].value,
                                    menu_title_right=row[2].value,
                                    belongs_to=belongs_to,
                                    defaults={'order': kedushot_order}
                                )
                                status = "Created" if created else "Found existing"
                                if verbose:
                                    self.stdout.write(
                                        f"{status} Kedushot: {last_kedushot.menu_title_left} in {belongs_to}")
                                kedushot_count += 1
                            kedushot_order += 1000
                        else:
                            if last_kedushot is None:
                                self.stdout.write(self.style.ERROR(
                                    f"Error: Trying to add menu item without a parent Kedushot. Row: {row[1].value}"
                                ))
                                continue

                            if dry_run:
                                if verbose:
                                    self.stdout.write(f"[DRY RUN] Would create MenuItem: {row[1].value}")
                            else:
                                menu_item, created = MenuItems.objects.get_or_create(
                                    menu_item=row[1].value,
                                    complement=row[2].value,
                                    defaults={'order': menu_order}
                                )
                                status = "Created" if created else "Found existing"
                                if verbose:
                                    self.stdout.write(f"{status} MenuItem: {menu_item.menu_item}")

                                # Add menu item to kedushot
                                last_kedushot.menu_items.add(menu_item)
                                menu_item_count += 1
                            menu_order += 1000
                    except Exception as e:
                        self.stdout.write(self.style.ERROR(f"Error processing row {row_count}: {e}"))
                        if not dry_run:
                            # If not in dry run mode, re-raise to trigger transaction rollback
                            raise

                if dry_run:
                    self.stdout.write(self.style.WARNING("DRY RUN - No changes were made to the database"))
                    transaction.set_rollback(True)
                else:
                    self.stdout.write(self.style.SUCCESS(
                        f"Successfully imported {row_count} rows of Kedushot data\n"
                        f"Created/Updated {kedushot_count} Kedushot entries and {menu_item_count} MenuItems"
                    ))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error: {e}"))
            return
