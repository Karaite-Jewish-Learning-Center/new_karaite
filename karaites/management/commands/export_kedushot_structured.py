from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from pathlib import Path
import json
import os


class Command(BaseCommand):
    help = 'Export the Index Format sheet from kedushot.xlsx to a structured JSON file'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output',
            type=str,
            default='kedushot_structured.json',
            help='Output JSON file path',
        )

    def handle(self, *args, **options):
        """Export the Index Format sheet to a structured JSON file"""
        output_file = options['output']

        # Check if Excel file exists
        excel_file = Path('karaites/management/commands/kedushot.xlsx')
        if not os.path.exists(excel_file):
            self.stdout.write(self.style.ERROR(f"Excel file not found at {excel_file}"))
            return

        self.stdout.write(self.style.SUCCESS(f"Found Excel file at {excel_file}"))

        try:
            wb = load_workbook(excel_file)

            # Check if the required sheet exists
            if 'Index Format' not in wb.sheetnames:
                self.stdout.write(self.style.ERROR("Sheet 'Index Format' not found in Excel file"))
                self.stdout.write(f"Available sheets: {wb.sheetnames}")
                return

            index_format_ws = wb['Index Format']
            self.stdout.write(self.style.SUCCESS("Found 'Index Format' sheet"))

            # Define categories
            weekly_poems = 'Poems for the Weekly Torah Portion'
            weekly_sabbath = 'Poems for the Weekly Sabbath'

            # Structure to hold our data
            structured_data = {
                weekly_poems: [],
                weekly_sabbath: []
            }

            current_kedushot = None
            current_category = None

            # Process rows starting from row 3 (as in your original script)
            for row_idx, row in enumerate(index_format_ws.iter_rows(min_row=3, values_only=True), 3):
                # Skip empty rows
                if row[1] is None:
                    continue

                # Extract values
                col_a = row[0]  # First column
                col_b = row[1]  # Second column
                col_c = row[2]  # Third column

                self.stdout.write(f"Processing row {row_idx}: {col_a}, {col_b}, {col_c}")

                # Check if this is a Kedushot header
                if col_a in ['Expand/Collapse', '#', 'A', 'B', 'C', 'D', 'E']:
                    # Determine category
                    category = weekly_sabbath if col_a in ['#', 'A', 'B', 'C', 'D', 'E'] else weekly_poems
                    current_category = category

                    # Create new Kedushot entry
                    current_kedushot = {
                        "menu_title_left": col_b,
                        "menu_title_right": col_c,
                        "belongs_to": category,
                        "order": (row_idx - 2) * 1000,  # Similar to your original logic
                        "menu_items": []
                    }

                    # Add to appropriate category
                    structured_data[category].append(current_kedushot)

                    self.stdout.write(f"Added Kedushot: {col_b} to category {category}")

                # Otherwise, it's a menu item
                elif current_kedushot is not None:
                    # Create menu item
                    menu_item = {
                        "menu_item": col_b,
                        "complement": col_c,
                        "order": (row_idx - 2) * 1000  # Similar to your original logic
                    }

                    # Add to current Kedushot
                    current_kedushot["menu_items"].append(menu_item)

                    self.stdout.write(f"Added MenuItem: {col_b} to Kedushot {current_kedushot['menu_title_left']}")
                else:
                    self.stdout.write(self.style.WARNING(
                        f"Warning: Row {row_idx} has a menu item without a parent Kedushot: {col_b}"
                    ))

            # Add summary information
            result = {
                "summary": {
                    "total_categories": len(structured_data),
                    "categories": {
                        category: {
                            "total_kedushot": len(kedushot_list),
                            "total_menu_items": sum(len(k["menu_items"]) for k in kedushot_list)
                        }
                        for category, kedushot_list in structured_data.items()
                    }
                },
                "data": structured_data
            }

            # Write to JSON file
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(result, f, indent=2, ensure_ascii=False)

            self.stdout.write(self.style.SUCCESS(
                f"Successfully exported structured data to {output_file}\n"
                f"Summary: {json.dumps(result['summary'], indent=2)}"
            ))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error: {e}"))
            return
