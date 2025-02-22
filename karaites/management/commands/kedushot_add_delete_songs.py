from django.core.management.base import BaseCommand, CommandError
from django.db.models.signals import post_save, post_delete
from ...models import Songs
from contextlib import contextmanager
from openpyxl import load_workbook
import argparse


@contextmanager
def disable_signals(signal, sender):
    """
        Disable signals for the given signal and sender
        So that cache is not cleared at every save or delete
    """
    # Store the original receivers
    receivers = signal.receivers
    signal.receivers = []
    try:
        yield
    finally:
        # Restore the original receivers
        signal.receivers = receivers


class Command(BaseCommand):
    help = 'Add or delete songs from the database'
    files = [
        "10_-_Piyut_-_Miketz.mp3", ''
        "11_-_Piyut_-_Vayigash.mp3",
        "12_-_Piyut_-_Vayechi.mp3",
        "13_-_Piyut_-_Shemot.mp3",
        "14_-_Piyut_-_Vaera.mp3",
        "15_-_Piyut_-_Bo_El_Paro.mp3",
        "16_-_Piyut_-_Beshalach.mp3",
        "17_-_Piyut_-_Yitro.mp3",
        "18_-_Piyut_-_Mishpatim.mp3",
        "19_-_Piyut_-_Teruma.mp3",
        "1_-_Piyut_-_Bereshit.mp3",
        "20_-_Piyut_-_Tetsave.mp3",
        "21_-_Piyut_-_Ki_Tisa.mp3",
        "22_-_Piyut_-_Vayakhel.mp3",
        "23_-_Piyut_-_Pekude.mp3",
        "24_-_Piyut_-_VaYikra.mp3",
        "25_-_Piyut_-_Tsav.mp3",
        "26_-_Piyut_-_Shemini.mp3",
        "27_-_Piyut_-_Tazria.mp3",
        "28_-_Piyut_-_Metzora.mp3",
        "29_-_Piyut_-_Achare_Mot.mp3",
        "2_-_Piyut_-_Noach.mp3",
        "30_-_Piyut_-_Kedoshim.mp3",
        "31_-_Piyut_-_Emor.mp3",
        "32_-_Piyut_-_Behar_Sinay.mp3",
        "33_-_Piyut_-_Bechukotay.mp3",
        "34_-_Piyut_-_Bemidbar_Sinay.mp3",
        "35_-_Piyut_-_Naso.mp3",
        "36_-_Piyut_-_Beha_alotecha.mp3",
        "37_-_Piyut_-_Shelach_Lecha.mp3",
        "38_-_Piyut_-_Korach.mp3",
        "39_-_Piyut_-_Chukat.mp3",
        "3_-_Piyut_-_Lech_Lecha.mp3",
        "40_-_Piyut_-_Balak.mp3",
        "41_-_Piyut_-_Pinechas.mp3",
        "42_-_Piyut_-_Matot.mp3",
        "43_-_Piyut_-_Masey.mp3",
        "44_-_Piyut_-_Devarim.mp3",
        "45-_Piyut_-_Vaetchanan.mp3",
        "46-_Piyut_-_Ekev.mp3",
        "47-_Piyut_-_Re_e.mp3",
        "48-_Piyut_-_Shofetim.mp3",
        "49_-_Piyut_-_Ki_Tetze.mp3",
        "4_-_Piyut_-_Vayera.mp3",
        "50_-_Piyut_-_Ki_Tavo.mp3",
        "51_-_Piyut_-_Nitzavim.mp3",
        "52_-_Piyut_-_Vayelech1.mp3",
        "52_-_Piyut_-_Vayelech2.mp3",
        "53_-_Piyut_-_Haazinu.mp3",
        "54_-_Piyut_-_Vezot_Haberacha.mp3",
        "5_-_Piyut_-_Chaye_Sarah.mp3",
        "6_-_Piyut_-_Toledot_Yitzchak.mp3",
        "7_-_Piyut_-_Vayetze_A.mp3",
        "7_-_Piyut_-_Vayetze_B.mp3",
        "8_-_Piyut_-_Vayishlach.mp3",
        "9_-_Piyut_-_Vayeshev.mp3"
    ]

    def add_arguments(self, parser):
        help = '''Add or delete songs from the database

Usage:
    python manage.py kedushot_add_delete_songs --add-songs
    python manage.py kedushot_add_delete_songs --delete-songs 
    python manage.py kedushot_add_delete_songs --xls-file <path to Excel file>
    python manage.py kedushot_add_delete_songs --delete-songs-xls <path to Excel file>

Example:
    ./manage.py kedushot_add_delete_songs --delete-songs-xls karaites/management/commands/kedushot.xlsx
'''

        # Add command line arguments using Django's parser
        parser.add_argument('--add-songs', action='store_true', help='Add songs to the database')
        parser.add_argument('--delete-songs', action='store_true', help='Delete songs from the database')
        parser.add_argument('--xls-file', type=str, help='Path to Excel file')
        parser.add_argument('--delete-songs-xls', type=str, help='Path to Excel file')

    def add_songs(self):
        # Disable both post_save and post_delete signals
        with disable_signals(post_save, Songs), disable_signals(post_delete, Songs):
            for file in self.files:
                song_title = file.split('-')[2].replace('.mp3', '').replace('_', '', 1).replace('_', ' ')
                song_file = 'songs/' + file
                song, created = Songs.objects.get_or_create(
                    song_title=song_title,
                    song_file=song_file
                )
                print(f"Processing {song.song_title}")

        print('End processing songs')

    def delete_songs(self):
        # Disable both post_save and post_delete signals
        with disable_signals(post_save, Songs), disable_signals(post_delete, Songs):
            for file in self.files:
                try:
                    song = Songs.objects.get(song_file=f'songs/{file}')
                    song.delete()
                    print(f"Deleting {song.song_title}")
                except Songs.DoesNotExist:
                    print(f"Song {file} not found")
        print('End deleting songs')

    def add_delete_songs_xls(self, xls_file, delete=False):
        # Disable both post_save and post_delete signals
        with disable_signals(post_save, Songs), disable_signals(post_delete, Songs):
            # read the xls file
            wb = load_workbook(xls_file)

            for sheet_name in wb.sheetnames:
                if sheet_name.startswith('Index Format') or sheet_name.startswith('DND') or sheet_name.startswith('Info'):
                    continue

                ws = wb[sheet_name]  # Get the worksheet by name
                # Skip header row and process all other rows
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]:  # Skip empty rows
                        continue

                    song_title = row[3]
                    possible_song_title = row[2]
                    if not song_title:
                        song_title = possible_song_title

                    song_file = row[0].replace(' ', '_').replace('.mp3', '') + '.mp3'
                    if not song_title:  # Skip rows without titles
                        continue

                    if delete:
                        try:
                            for song in Songs.objects.filter(song_file=f'songs/{song_file}'):
                                song.delete()
                                print(f"Deleting {song.song_title}")
                        except Songs.DoesNotExist:
                            print(f"Song {song_file} not found")

                    else:
                        song, created = Songs.objects.get_or_create(
                            song_title=song_title,
                            song_file=f'songs/{song_file}'
                        )
                        status = "Created" if created else "Already exists"
                        print(f"{status}: {song.song_title} ({song.song_file})")

        print('End processing songs')

    def handle(self, *args, **options):
        # Fix reference to self.files in add_songs method
        if options['add_songs']:
            self.add_songs()
        if options['delete_songs']:
            self.delete_songs()
        if options.get('xls_file'):
            self.add_delete_songs_xls(options['xls_file'])
        if options.get('delete_songs_xls'):
            self.add_delete_songs_xls(options['delete_songs_xls'], delete=True)
