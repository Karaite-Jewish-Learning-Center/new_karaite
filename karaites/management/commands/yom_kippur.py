import sys
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from torchgen.selective_build.operator import strip_operator_overload_name

from ...models import (Songs,
                       FirstLevel,
                       Classification,
                       KaraitesBookAsArray,
                       KaraitesBookDetails,
                       FILLER)
from ...utils import (Stack,
                      convert_time_string)

from django.core.files import File
from pathlib import Path


class Command(BaseCommand):

    @staticmethod
    def save_data(liturgy_details, songs, line, line_number):
        # save liturgyBook
        try:
            liturgy_book = KaraitesBookAsArray.objects.get(book=liturgy_details, song=songs, line_number=line_number)
        except KaraitesBookAsArray.DoesNotExist:
            liturgy_book = KaraitesBookAsArray()

        liturgy_book.book = liturgy_details
        liturgy_book.song = songs
        liturgy_book.book_text = line
        liturgy_book.line_number = line_number
        liturgy_book.better_book = True
        liturgy_book.save()

    @staticmethod
    def save_song(english_name, song_file, path):
        songs = Songs.objects.get(song_title=english_name)
        try:
            songs = Songs.objects.get(song_title=english_name)
        except Songs.DoesNotExist:
            songs = Songs()

        # try:
        #     song_file_name = path / song_file
        #     songs.song_title = english_name
        #     songs.song_file.save(song_file, File(open(song_file_name, 'rb')))
        #     songs.save()
        # except FileNotFoundError:
        #     return None

        return songs

    def add_arguments(self, parser):
        parser.add_argument(
            '--xls_file',
            default='Kedushot and Piyyut Parasha.xlsx',
            help='Update LiturgyBooks table with with excel data file.',
        )

    def handle(self, *args, **options):
        """ Read excel import songs, book text and audio, markers """

        books = [
            'Anenu Adonai Anenu (Yom Kippur).xlsx',
            'Anenu El Hai Vekayyam.xlsx',
            "Ase Lema'an Shimcha Haggadol.xlsx",
            'Berogez Rahem Tizkor (yom kippur).xlsx',
            'El Rahum Shemach (Yom Kippur).xlsx',
            'Elohē yisra’el al tishkaḥ yisra’el.xlsx'
        ]
        path = Path() / 'data_karaites/HTML/Liturgy/Yom Kippur/'
        order = 1000
        for book in books:
            songs = None
            file_name = path / book

            wb = load_workbook(file_name)
            stack = Stack()

            print('Processing book: ', book)
            # print all sheet names
            # print(wb.sheetnames)
            ws = wb['Text']

            # song details
            hebrew_name = ws['C2'].value
            english_name = ws['D2'].value
            song_file = ws['A2'].value
            has_times = isinstance(ws['P2'].value, (float, int))

            if song_file is not None:
               song_file = song_file.replace('.', '')

            print('Book: ', book, ' Hebrew name: ', hebrew_name, ' English name: ', english_name, ' song file: ', song_file)

            KaraitesBookDetails.objects.filter(book_title_en=english_name).delete()

            liturgy_details = KaraitesBookDetails()
            liturgy_details.first_level = FirstLevel.objects.get(first_level='Prayers & Songs')
            liturgy_details.book_classification = Classification.objects.get(
                classification_name='Yom Kippur')
            liturgy_details.occasion = ws['B2'].value
            liturgy_details.book_title_he = hebrew_name
            liturgy_details.book_title_en = english_name
            liturgy_details.better_book = True
            liturgy_details.better_intro = []
            liturgy_details.better_toc = []
            liturgy_details.book_language = 'he-en'
            liturgy_details.author = None
            liturgy_details.order = order
            liturgy_details.published = True
            liturgy_details.save()
            order += 1000

            if has_times and song_file is not None:
                song = self.save_song(english_name, song_file, path)
                liturgy_details.songs.add(song)

            row = 2
            line_number = 0
            spreadsheet_line = 1
            english_translation = []
            # use some empty lines on top to better display the text on the grid
            filler = FILLER
            # hebrew_text = [filler, filler, filler, filler]
            hebrew_text = []
            # audio_start
            if ws[f'O{row}'].value is not None:
                stack.push(float(ws[f'O{row}'].value))
            else:
                stack.push(None)

            songs_id = -1
            while True:
                # line number
                if ws[f'I{row}'].value is None:
                    break

                if has_times:
                    # maybe more than one file song per book
                    if ws[f'A{row}'].value is not None:
                        song_file = ws[f'A{row}'].value
                        if song_file is not None:
                            song_file = song_file.replace('.', '')

                        songs = self.save_song(english_name, song_file, path)
                        print('Saving song: ', song_file, ' for book: ', english_name)
                        input('Press Enter to continue...')
                        if songs is not None:
                            songs_id = songs.id

                audio_start = stack.pop()

                # may be None, a string or a float
                value = ws[f'P{row}'].value
                if value is not None:
                    value = str(ws[f'P{row}'].value)
                print("value: ", value)
                stack.push(value)

                # [hebrew, transliteration, english audio_start, audio_end, song_id, reciter, censored, line_number, comments, end of verse, section or subtext, filler, song end, Arabic]
                # print("hebrew: ", ws[f'J{row}'].value)
                # print("transliteration: ", ws[f'K{row}'].value)
                # print("english: ", ws[f'L{row}'].value)
                # print("audio_start: ", audio_start)
                # print("audio_end: ", stack.peek())
                # print("song_id: ", songs_id)
                # print("reciter: ", ws[f'H{row}'].value)
                # print("censored: ", ws[f'G{row}'].value)
                # print("line_number: ", ws[f'I{row}'].value)
                # print("comments: ", ws[f'M{row}'].value)
                # print("end of verse, section or subtext: ", ws[f'F{row}'].value)
                # print("---------------------------------------------------")
                # input('Press Enter to continue...')
                hebrew_text.append([
                    ws[f'J{row}'].value,  # hebrew
                    ws[f'K{row}'].value,  # transliteration
                    ws[f'L{row}'].value,  # english
                    audio_start,
                    stack.peek(),  # audio_end
                    songs_id,  # song_id
                    ws[f'H{row}'].value,  # reciter
                    ws[f'G{row}'].value,  # censored
                    ws[f'I{row}'].value,  # line_number
                    ws[f'M{row}'].value,  # comments
                    0,  # end of verse, section or subtext? No
                    0,  # filler
                    0,  # song end
                    "", # Comments
                    "0",
                    "0",
                    "0",
                    "0",
                    "0",
                    "0",
                    "0",
                ])

                english_translation.append(
                    ['', '', ws[f'L{row}'].value, '', '', songs_id, '', '', '', '', 0, 1, 0, '0', '0', '0', '0', '0', '0', '0'])

                # end of verse, section or subtext
                end = ws[f'F{row}'].value
                if end is not None and ws[f'F{row}'].value.find('end') >= 0:
                    # save hebrew text
                    hebrew_text[-1][10] = 1  # end of verse, subtext
                    if end.find('<end subtext>') >= 0:
                        hebrew_text[-1][12] = 1  # end of song

                    for hebrew in hebrew_text:
                        self.save_data(liturgy_details, songs, hebrew, line_number)
                        line_number += 1

                    # save english translation
                    english_translation[-1][10] = 1  # end of verse, section or subtext? Yes

                    for english in english_translation:
                        self.save_data(liturgy_details, songs, english, line_number)
                        line_number += 1

                    english_translation = []
                    hebrew_text = []

                print('Processing  book: ', book, ' song: ', english_name, ' line_number: ', spreadsheet_line, )
                spreadsheet_line += 1
                row += 1

            # append some empty lines at bottom to better display the text on the grid
            for _ in range(10):
                self.save_data(liturgy_details, songs, filler, line_number)
                line_number += 1
