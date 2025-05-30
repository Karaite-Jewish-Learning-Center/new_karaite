import re
from django.conf import settings
from django.core.management.base import BaseCommand
from django.core.management import call_command
from openpyxl import load_workbook
from pathlib import Path
from contextlib import contextmanager
from django.contrib.auth.models import User
from ...models import (Author,
                       Songs,
                       FirstLevel,
                       Classification,
                       KaraitesBookAsArray,
                       KaraitesBookDetails,
                       FILLER)
from ...utils import convert_time_string
from django.db.models.signals import post_save, post_delete


@contextmanager
def disable_signals(signal, sender):
    """
        Disable signals for the given signal and sender
        So that cache is not cleared at every save or delete
    """
    # Get all receivers for this signal and sender
    receivers = signal.receivers[:]
    # Disconnect all receivers
    signal.receivers = []
    try:
        yield
    finally:
        # Restore original receivers
        signal.receivers = receivers


class Command(BaseCommand):
    help = 'Fill Database details from sheets'

    def handle(self, *args, **options):
        """Import Kedushot and Piyyutim La-Parashiyyot sheets"""
        File_Name = 0
        Pattern = 1
        Hebrew_Name = 2
        English_Name = 3
        In_Place_of = 4
        Reciter = 5
        Hebrew_Line = 6
        Hebrew_Text = 7
        English_Transliteration = 8
        English_Translation = 9
        Comments = 10
        Time_Starting = 11
        Time_Ending = 12

        # Wrap the entire operation in the disable_signals context manager
        with disable_signals(post_save, KaraitesBookDetails), \
                disable_signals(post_delete, KaraitesBookDetails), \
                disable_signals(post_save, KaraitesBookAsArray), \
                disable_signals(post_delete, KaraitesBookAsArray):

            # call_command('kedushot_add_songs')
            open_file = Path('karaites/management/commands/kedushot.xlsx')
            wb = load_workbook(open_file)
            law = FirstLevel.objects.get(first_level='Prayers & Songs')
            classification = Classification.objects.get(classification_name='Kedushot and Piyyutim La-Parashiyyot')
            book_language = 'he-en'
            user = User.objects.get(username='System')
            # open and process each sheet in the workbook
            for sheet_name in wb.sheetnames:
                # Skip unwanted sheets
                if sheet_name.startswith('Index Format') or sheet_name.startswith('DND') or sheet_name.startswith('Info'):
                    continue

                # Get the worksheet
                ws = wb[sheet_name]

                print('Processing sheet', sheet_name)
                # read the second  row
                for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                    # Unpack row values
                    song_file = row[File_Name].replace(' ', '_').replace('.mp3', '')+'.mp3'
                    pattern = row[Pattern] or ''
                    hebrew_name = row[Hebrew_Name] or ''
                    english_name = row[English_Name] or ''
                    inplace_of = row[In_Place_of]
                    reciter = row[Reciter]
                    hebrew_line = row[Hebrew_Line]
                    hebrew_text = row[Hebrew_Text]
                    english_transliteration = row[English_Transliteration]
                    english_translation = row[English_Translation]
                    comments = row[Comments]
                    audio_start = convert_time_string(row[Time_Starting] or 0)
                    audio_end = convert_time_string(row[Time_Ending] or 0)

                    if english_name == '':
                        english_name = hebrew_name
                    english_name = english_name.strip()
                    # if settings.DEBUG:
                    #     print('Song Name', song_file)
                    #     print('Pattern', pattern)
                    #     print('Hebrew Name', hebrew_name)
                    #     print('English Name', english_name)
                    #     print('Inplace Of', inplace_of)
                    #     print('Reciter', reciter)
                    #     print('Hebrew Line', hebrew_line)
                    #     print('Hebrew Text', hebrew_text)
                    #     print('English Transliteration', english_transliteration)
                    #     print('English Translation', english_translation)
                    #     print('Comments', comments)
                    #     print('Audio End', audio_end)

                    # author is in the name of thr song preceded by "by"
                    author_name = re.search(r'by\s+(\w+)', english_name)
                    if author_name:
                        try:
                            author = Author.objects.get(name=author_name.group(1))
                        except Author.DoesNotExist:
                            author = Author.objects.create(name=author_name.group(1))
                    else:
                        author = Author.objects.get(name='Aaron ben Joseph')

                    # create details instance
                    instance, created = KaraitesBookDetails.objects.get_or_create(
                        first_level=law,
                        book_classification=classification,
                        book_language=book_language,
                        author=author,
                        book_title_en=english_name,
                        book_title_he=hebrew_name,
                        user=user,
                        better_book=True,
                        better_intro=[],
                        better_toc=[],
                    )
                    print(f'song title:"{english_name}"')
                    print('song_file', song_file)

                    song = Songs.objects.get(song_title=english_name)
                    try:
                        song = Songs.objects.get(song_title=english_name)
                    except Songs.DoesNotExist:
                        print(f'Song {english_name} not found')
                        song = None

                    if song is not None:
                        instance.songs.add(song)
                        instance.save()

                    censored = 0
                    break_point = 0
                    song_ends = 0
                    line_number = 0
                    # read rest of the rows
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                        if row[7] is None:
                            break

                        reciter = row[Reciter]
                        hebrew_line = row[Hebrew_Line]
                        hebrew_text = row[Hebrew_Text]
                        english_transliteration = row[English_Transliteration]
                        english_translation = row[English_Translation]
                        comments = row[Comments]
                        audio_end = convert_time_string(row[Time_Ending])
                        # [hebrew, transliteration, english, audio_start, audio_end, song_id, reciter, censored, line_number, break, song end, comments, pattern, reserved, reserved, reserved, reserved, reserved    ]
                        KaraitesBookAsArray.objects.get_or_create(
                            book=instance,
                            song=song,
                            book_text=[hebrew_text,
                                       english_transliteration,
                                       english_translation,
                                       audio_start,
                                       audio_end,
                                       song.id if song else None,
                                       reciter,
                                       censored,
                                       hebrew_line,
                                       break_point,
                                       song_ends,
                                       comments,
                                       pattern,
                                       0,
                                       0,
                                       0,
                                       0,
                                       0,
                                       ],
                            line_number=line_number,
                        )
                        line_number += 1
                        audio_start = audio_end
                        add_filler = False
                        # patterns ['Continuous, until Hazzan', 'Quatrain', 'A', 'Couplet']
                        if pattern.startswith('Continuous') or pattern in ['A', '',] and reciter == 'Hazzan':
                            add_filler = True
                        elif pattern == 'Quatrain' and reciter == 'Hazzan':
                            add_filler = True
                        elif pattern == 'Couplet' and (reciter == 'Congregant' or reciter == 'Hazzan'):
                            add_filler = True

                        if add_filler:
                            for i in range(0, add_filler):
                                KaraitesBookAsArray.objects.get_or_create(
                                    book=instance,
                                    song=song,
                                    book_text=FILLER,
                                    line_number=line_number,
                                )
                                line_number += 1

            # update the KaraitesBookDetails order field
            # based on the Index Format sheet

            order = 1_000
            misses = 0
            index_sheet = wb['Index Format']

            for row in index_sheet.iter_rows(min_row=1, values_only=True):
                if not row[0] or row[0] in ['#', 'Expand/Collapse'] or not row[2]:
                    continue

                book_title = row[1].strip()
                try:
                    print("Try - 1:", book_title)
                    book_details = KaraitesBookDetails.objects.get(
                        book_title_en__startswith=book_title, published=False)
                    book_details.order = order
                    book_details.kedushot_order = f'{row[0]}'.strip()
                    book_details.save()
                    self.stdout.write(f"Updated order for '{book_title}' to {order}")

                except (KaraitesBookDetails.DoesNotExist, KaraitesBookDetails.MultipleObjectsReturned):
                    misses += 1
                    self.stdout.write(self.style.WARNING(f"Book '{book_title}' not found"))

                order += 1000

            if misses > 0:
                self.stdout.write(self.style.WARNING(f"Missed {misses} books"))
