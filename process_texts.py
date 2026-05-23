#!/usr/bin/env python3
"""
Process Karaite texts from HTML/XLSX source files into clean JSON format.
This creates a simple, forkable data structure for the static site.
"""

import json
import os
import re
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# Source paths
SOURCE_HTML = Path("/Users/shawn/knowledge-chatbot/documents/new_karaite/data_karaites/HTML")
SOURCE_XLSX = Path("/Users/shawn/knowledge-chatbot/documents/new_karaite/data_karaites/Out_xls")
OUTPUT_DIR = Path("/Users/shawn/karaite-texts/data/texts")

def slugify(text):
    """Convert text to URL-friendly slug."""
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_-]+', '-', text)
    return text

def clean_whitespace(text):
    """Clean up whitespace artifacts from HTML parsing."""
    if not text:
        return ""
    # Replace newlines and multiple spaces with single space
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def get_text_with_spacing(element):
    """Extract text from element, preserving spacing between paragraphs."""
    if element is None:
        return ""
    
    # Get all paragraph elements
    paragraphs = element.find_all('p')
    if paragraphs:
        # Join paragraph texts with / separator
        texts = []
        for p in paragraphs:
            text = clean_whitespace(p.get_text())
            if text:
                texts.append(text)
        return ' / '.join(texts)  # Use / as line separator for display
    else:
        return clean_whitespace(element.get_text())

def extract_metadata_from_html(soup):
    """Extract metadata fields from HTML content."""
    metadata = {}
    text = soup.get_text()
    
    # Common metadata patterns
    patterns = {
        'category': r'Category:\s*(.+?)(?:\n|$)',
        'genre': r'Genre:\s*(.+?)(?:\n|$)',
        'occasion': r'Occasion:\s*(.+?)(?:\n|$)',
        'composer': r'Composer:\s*(.+?)(?:\n|$)',
        'location': r'Location:\s*(.+?)(?:\n|$)',
        'date': r'Date:\s*(.+?)(?:\n|$)',
        'acrostic': r'Acrostic:\s*(.+?)(?:\n|$)',
        'source': r'Source:\s*(.+?)(?:\n|$)',
        'meter': r'Description of Meter:\s*(.+?)(?:\n|$)',
        'davidson': r'Davidson number:\s*(.+?)(?:\n|$)',
        'karaite_origin': r'Karaite origin:\s*(.+?)(?:\n|$)',
    }
    
    for key, pattern in patterns.items():
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            metadata[key] = match.group(1).strip()
    
    return metadata

def extract_introduction(soup):
    """Extract introduction text from HTML."""
    text = soup.get_text()
    
    # Find introduction section
    intro_match = re.search(r'Introduction:\s*(.+?)(?:Category:|Genre:|$)', text, re.DOTALL | re.IGNORECASE)
    if intro_match:
        intro = intro_match.group(1).strip()
        # Clean up whitespace
        intro = re.sub(r'\s+', ' ', intro)
        return intro
    return ""

def extract_about_author(soup):
    """Extract 'About the Author' section."""
    text = soup.get_text()
    
    match = re.search(r'About the Author:\s*(.+?)(?:Sources:|$)', text, re.DOTALL | re.IGNORECASE)
    if match:
        about = match.group(1).strip()
        about = re.sub(r'\s+', ' ', about)
        return about
    return ""

def has_hebrew(text):
    """Check if text contains Hebrew characters."""
    return bool(re.search(r'[\u0590-\u05FF]', text))

def extract_table_content(soup):
    """Extract content from HTML tables (liturgical texts format).
    
    Table structure:
    - Odd rows (1,3,5...): Two cells - Transliteration (left) | Hebrew (right)
    - Even rows (2,4,6...): One cell with colspan=2 - English translation
    """
    content = []
    
    tables = soup.find_all('table')
    for table in tables:
        rows = table.find_all('tr')
        
        i = 0
        while i < len(rows):
            row = rows[i]
            cells = row.find_all('td')
            
            # Check for colspan (English translation row)
            if len(cells) == 1 or (cells and cells[0].get('colspan')):
                # This might be an English-only row - skip, we'll get it with the Hebrew row
                i += 1
                continue
            
            if len(cells) >= 2:
                # Two-cell row: should be Transliteration | Hebrew
                hebrew = ""
                translit = ""
                english = ""
                
                # Process each cell
                for cell in cells:
                    cell_text = get_text_with_spacing(cell)
                    if not cell_text:
                        continue
                    
                    # Check direction attribute or Hebrew characters
                    is_rtl = cell.get('dir') == 'rtl' or 'direction:rtl' in str(cell.get('style', ''))
                    cell_has_hebrew = has_hebrew(cell_text)
                    
                    if cell_has_hebrew or is_rtl:
                        hebrew = cell_text
                    else:
                        translit = cell_text
                
                # Look for English in the next row (colspan row)
                if i + 1 < len(rows):
                    next_row = rows[i + 1]
                    next_cells = next_row.find_all('td')
                    if next_cells:
                        # Check if it's a colspan row
                        first_cell = next_cells[0]
                        colspan = first_cell.get('colspan')
                        if colspan or len(next_cells) == 1:
                            english = get_text_with_spacing(next_row)
                            # Skip the English row
                            i += 1
                
                # Only add if we have actual content
                if hebrew or translit:
                    content.append({
                        "hebrew": hebrew,
                        "transliteration": translit,
                        "english": english
                    })
            
            i += 1
    
    return content

def process_liturgy_html(filepath):
    """Process a liturgical HTML file into JSON format."""
    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        html = f.read()
    
    soup = BeautifulSoup(html, 'html.parser')
    
    # Extract titles
    title_en = ""
    title_he = ""
    
    # Look for centered paragraphs at the start (titles)
    for p in soup.find_all('p')[:10]:
        style = p.get('style', '') or ''
        align = p.get('align', '')
        if 'center' in style or align == 'center':
            text = p.get_text(strip=True)
            if has_hebrew(text) and not title_he:
                title_he = text
            elif text and not title_en and len(text) < 100 and not has_hebrew(text):
                title_en = text
    
    # Determine category from path
    parts = filepath.relative_to(SOURCE_HTML).parts
    category = parts[0] if len(parts) > 1 else ""
    subcategory = parts[1] if len(parts) > 2 else (parts[0] if len(parts) > 1 else "")
    
    # For deeply nested paths, use the folder name
    if len(parts) > 2:
        subcategory = parts[-2]  # Parent folder of the file
    
    # Build the JSON structure
    result = {
        "id": slugify(filepath.stem),
        "title_en": title_en or filepath.stem,
        "title_he": title_he,
        "category": category,
        "subcategory": subcategory,
        "introduction": extract_introduction(soup),
        "about_author": extract_about_author(soup),
        "metadata": extract_metadata_from_html(soup),
        "content": extract_table_content(soup),
        "source_file": str(filepath.name)
    }
    
    return result

def process_xlsx(filepath):
    """Process an Excel file into JSON format."""
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        
        content = []
        title_en = filepath.stem
        title_he = ""
        
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not any(row):
                continue
            
            # Try to identify Hebrew, transliteration, English columns
            hebrew = ""
            translit = ""
            english = ""
            
            for cell in row:
                if cell is None:
                    continue
                text = str(cell).strip()
                if not text:
                    continue
                
                if has_hebrew(text):
                    if not title_he and len(text) < 50:
                        title_he = text
                    hebrew = text
                elif re.search(r'^[A-Za-z]', text):
                    if not translit:
                        translit = text
                    elif not english:
                        english = text
            
            if hebrew or translit:
                content.append({
                    "hebrew": hebrew,
                    "transliteration": translit,
                    "english": english
                })
        
        return {
            "id": slugify(filepath.stem),
            "title_en": title_en,
            "title_he": title_he,
            "category": "Liturgy",
            "subcategory": filepath.parent.name,
            "content": content,
            "source_file": str(filepath.name)
        }
    except Exception as e:
        print(f"Error processing {filepath}: {e}")
        return None

def build_catalog(texts):
    """Build the catalog.json index file."""
    catalog = {
        "categories": {},
        "texts": []
    }
    
    for text in texts:
        # Add to texts list
        catalog["texts"].append({
            "id": text["id"],
            "title_en": text["title_en"],
            "title_he": text["title_he"],
            "category": text["category"],
            "subcategory": text.get("subcategory", ""),
            "has_audio": text.get("audio") is not None
        })
        
        # Organize by category
        cat = text["category"]
        subcat = text.get("subcategory", "")
        
        if cat not in catalog["categories"]:
            catalog["categories"][cat] = {"subcategories": {}, "texts": []}
        
        if subcat and subcat != cat:
            if subcat not in catalog["categories"][cat]["subcategories"]:
                catalog["categories"][cat]["subcategories"][subcat] = []
            catalog["categories"][cat]["subcategories"][subcat].append(text["id"])
        else:
            catalog["categories"][cat]["texts"].append(text["id"])
    
    return catalog

def main():
    """Main processing function."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    processed_texts = []
    
    # Process HTML files
    print("Processing HTML files...")
    for html_file in SOURCE_HTML.rglob("*.html"):
        # Skip scratch folder
        if 'scratch' in str(html_file):
            continue
            
        print(f"  {html_file.name}")
        try:
            text_data = process_liturgy_html(html_file)
            if text_data and (text_data["content"] or text_data["introduction"]):
                # Save individual text file
                output_file = OUTPUT_DIR / f"{text_data['id']}.json"
                with open(output_file, 'w', encoding='utf-8') as f:
                    json.dump(text_data, f, ensure_ascii=False, indent=2)
                processed_texts.append(text_data)
        except Exception as e:
            print(f"    Error: {e}")
    
    # Process XLSX files (for texts without HTML)
    print("\nProcessing XLSX files...")
    existing_ids = {t["id"] for t in processed_texts}
    for xlsx_file in SOURCE_XLSX.rglob("*.xlsx"):
        slug = slugify(xlsx_file.stem)
        if slug not in existing_ids:
            print(f"  {xlsx_file.name}")
            try:
                text_data = process_xlsx(xlsx_file)
                if text_data and text_data["content"]:
                    output_file = OUTPUT_DIR / f"{text_data['id']}.json"
                    with open(output_file, 'w', encoding='utf-8') as f:
                        json.dump(text_data, f, ensure_ascii=False, indent=2)
                    processed_texts.append(text_data)
            except Exception as e:
                print(f"    Error: {e}")
    
    # Build and save catalog
    print("\nBuilding catalog...")
    catalog = build_catalog(processed_texts)
    with open(OUTPUT_DIR.parent / "catalog.json", 'w', encoding='utf-8') as f:
        json.dump(catalog, f, ensure_ascii=False, indent=2)
    
    print(f"\nDone! Processed {len(processed_texts)} texts.")
    print(f"Output: {OUTPUT_DIR}")

if __name__ == "__main__":
    main()
