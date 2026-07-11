#!/usr/bin/env python3
"""
Build a beautiful static website from all XLSX sources.
Creates JSON data files and generates the site structure.
"""

import json
import re
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook

# Paths
BASE_DIR = Path("/Users/shawn/karaite-texts")
OUTPUT_DIR = BASE_DIR / "site"
DATA_DIR = OUTPUT_DIR / "data"

SOURCES = {
    'kedushot': Path("/Users/shawn/knowledge-chatbot/documents/new_karaite/karaites/management/commands/kedushot.xlsx"),
    'out_xls': Path("/Users/shawn/knowledge-chatbot/documents/new_karaite/data_karaites/Out_xls"),
    'html_xlsx': Path("/Users/shawn/knowledge-chatbot/documents/new_karaite/data_karaites/HTML"),
    'word_xlsx': Path("/Users/shawn/knowledge-chatbot/documents/new_karaite/data_karaites/Word Documents"),
    'converted_html': BASE_DIR / "xlsx_from_html",
    'converted_xml': BASE_DIR / "xlsx_from_xml",
}


def slugify(text):
    """Convert text to URL-safe slug."""
    text = re.sub(r'[^\w\s-]', '', text.lower())
    return re.sub(r'[-\s]+', '-', text).strip('-')


def clean_text(text):
    """Clean text for JSON output."""
    if not text:
        return ""
    return str(text).strip()


def parse_formatting(text):
    """Parse formatting markers into structured format."""
    if not text:
        return {"raw": "", "formatted": []}
    
    text = str(text)
    formatted = []
    
    # Simple approach: return raw text with markers
    # The frontend will parse these
    return {
        "raw": text,
        "has_formatting": any(m in text for m in ['_', '**', '{{', '['])
    }


def determine_category(path, name):
    """Determine category from file path and name."""
    path_str = str(path).lower()
    name_lower = name.lower()
    
    if 'liturgy' in path_str or 'yom kippur' in path_str:
        if 'havdala' in path_str:
            return 'Liturgy', 'Havdala Songs'
        elif 'shabbat' in path_str:
            return 'Liturgy', 'Shabbat Songs'
        elif 'yom kippur' in path_str or 'kippur' in name_lower:
            return 'Liturgy', 'Yom Kippur'
        elif 'kedushot' in path_str:
            return 'Liturgy', 'Kedushot'
        else:
            return 'Liturgy', 'General'
    elif 'halakh' in path_str:
        return 'Halakhah', None
    elif 'poetry' in path_str:
        return 'Poetry', None
    elif 'polemic' in path_str:
        return 'Polemics', None
    elif 'comment' in path_str:
        return 'Comments', None
    elif 'exhort' in path_str:
        return 'Exhortatory', None
    else:
        return 'General', None


def parse_kedushot_xlsx():
    """Parse the kedushot.xlsx master file."""
    texts = []
    path = SOURCES['kedushot']
    
    wb = load_workbook(path, read_only=True, data_only=True)
    
    for sheet_name in wb.sheetnames:
        if sheet_name in ['Index Format', 'Info', 'DND']:
            continue
        
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        if len(rows) < 2:
            continue
        
        headers = rows[0]
        
        # Find column indices
        col_map = {}
        for i, h in enumerate(headers):
            if h:
                h_lower = str(h).lower()
                if 'hebrew text' in h_lower or h_lower == 'hebrew text':
                    col_map['hebrew'] = i
                elif 'transliteration' in h_lower:
                    col_map['transliteration'] = i
                elif 'english translation' in h_lower or 'translation' in h_lower:
                    col_map['english'] = i
                elif 'comment' in h_lower:
                    col_map['comments'] = i
                elif 'hebrew name' in h_lower:
                    col_map['title_he'] = i
                elif 'english name' in h_lower:
                    col_map['title_en'] = i
        
        # Extract content
        content = []
        title_en = sheet_name
        title_he = ""
        
        for row in rows[1:]:
            if not row or all(c is None for c in row):
                continue
            
            # Get title from first data row
            if not title_he and 'title_he' in col_map and row[col_map['title_he']]:
                title_he = clean_text(row[col_map['title_he']])
            if 'title_en' in col_map and row[col_map['title_en']]:
                title_en = clean_text(row[col_map['title_en']])
            
            hebrew = clean_text(row[col_map.get('hebrew', 7)] if col_map.get('hebrew') is not None else (row[7] if len(row) > 7 else ''))
            translit = clean_text(row[col_map.get('transliteration', 8)] if col_map.get('transliteration') is not None else (row[8] if len(row) > 8 else ''))
            english = clean_text(row[col_map.get('english', 9)] if col_map.get('english') is not None else (row[9] if len(row) > 9 else ''))
            comments = clean_text(row[col_map.get('comments', 10)] if col_map.get('comments') is not None else (row[10] if len(row) > 10 else ''))
            
            if hebrew or translit or english:
                content.append({
                    'hebrew': hebrew,
                    'transliteration': translit,
                    'english': english,
                    'comments': comments
                })
        
        if content:
            # Determine subcategory for kedushot
            if sheet_name.startswith(('A.', 'B.', 'C.', 'D.', 'E.')):
                subcategory = 'Weekly Kedushot'
            else:
                subcategory = 'Torah Portions'
            
            texts.append({
                'id': slugify(sheet_name),
                'title_en': title_en,
                'title_he': title_he,
                'category': 'Liturgy',
                'subcategory': subcategory,
                'source': 'kedushot.xlsx',
                'content': content
            })
    
    wb.close()
    return texts


def parse_xlsx_file(path, sheet_name=None):
    """Parse a single XLSX file."""
    wb = load_workbook(path, read_only=True, data_only=True)
    
    texts = []
    sheets_to_process = [sheet_name] if sheet_name else wb.sheetnames
    
    for sname in sheets_to_process:
        if sname not in wb.sheetnames:
            continue
        if sname.lower() in ['metadata', 'footnotes', 'alignment', 'introduction', 'toc']:
            continue
        
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        
        if len(rows) < 2:
            continue
        
        headers = rows[0]
        
        # Find column indices
        col_map = {}
        for i, h in enumerate(headers or []):
            if h:
                h_lower = str(h).lower()
                if 'hebrew text' in h_lower:
                    col_map['hebrew'] = i
                elif 'transliteration' in h_lower:
                    col_map['transliteration'] = i
                elif 'english translation' in h_lower or 'translation' in h_lower:
                    col_map['english'] = i
                elif 'comment' in h_lower:
                    col_map['comments'] = i
                elif 'arabic' in h_lower:
                    col_map['arabic'] = i
        
        # Default columns if not found
        if 'hebrew' not in col_map:
            col_map['hebrew'] = 7
        if 'transliteration' not in col_map:
            col_map['transliteration'] = 8
        if 'english' not in col_map:
            col_map['english'] = 9
        if 'comments' not in col_map:
            col_map['comments'] = 10
        
        # Extract content
        content = []
        for row in rows[1:]:
            if not row or all(c is None for c in row):
                continue
            
            def get_col(key):
                idx = col_map.get(key)
                if idx is not None and len(row) > idx:
                    return clean_text(row[idx])
                return ""
            
            hebrew = get_col('hebrew')
            translit = get_col('transliteration')
            english = get_col('english')
            comments = get_col('comments')
            arabic = get_col('arabic')
            
            if hebrew or translit or english or arabic:
                entry = {
                    'hebrew': hebrew,
                    'transliteration': translit,
                    'english': english,
                }
                if comments:
                    entry['comments'] = comments
                if arabic:
                    entry['arabic'] = arabic
                content.append(entry)
        
        if content:
            # Try to get metadata
            title_en = path.stem
            title_he = ""
            intro = ""
            
            if 'Metadata' in wb.sheetnames:
                meta_ws = wb['Metadata']
                for row in meta_ws.iter_rows(values_only=True):
                    if row and row[0]:
                        field = str(row[0]).lower()
                        value = clean_text(row[1]) if len(row) > 1 else ""
                        if 'english' in field and 'title' in field:
                            title_en = value or title_en
                        elif 'hebrew' in field and 'title' in field:
                            title_he = value
            
            if 'Introduction' in wb.sheetnames:
                intro_ws = wb['Introduction']
                intro_parts = []
                for row in intro_ws.iter_rows(values_only=True):
                    if row and row[0]:
                        intro_parts.append(clean_text(row[0]))
                intro = '\n\n'.join(intro_parts)
            
            category, subcategory = determine_category(path, title_en)
            
            text_data = {
                'id': slugify(title_en),
                'title_en': title_en,
                'title_he': title_he,
                'category': category,
                'source': path.name,
                'content': content
            }
            
            if subcategory:
                text_data['subcategory'] = subcategory
            if intro:
                text_data['introduction'] = intro
            
            texts.append(text_data)
    
    wb.close()
    return texts


def collect_all_texts():
    """Collect all texts from all sources."""
    all_texts = []
    seen_ids = set()
    
    # 1. Kedushot
    print("Processing kedushot.xlsx...")
    kedushot_texts = parse_kedushot_xlsx()
    for t in kedushot_texts:
        if t['id'] not in seen_ids:
            all_texts.append(t)
            seen_ids.add(t['id'])
    print(f"  Found {len(kedushot_texts)} texts")
    
    # 2. Out_xls directory
    print("Processing Out_xls...")
    count = 0
    for xlsx in SOURCES['out_xls'].rglob("*.xlsx"):
        if 'scratch' in str(xlsx).lower() or xlsx.stem.startswith('~'):
            continue
        try:
            texts = parse_xlsx_file(xlsx)
            for t in texts:
                if t['id'] not in seen_ids:
                    all_texts.append(t)
                    seen_ids.add(t['id'])
                    count += 1
        except Exception as e:
            print(f"  Error processing {xlsx.name}: {e}")
    print(f"  Found {count} texts")
    
    # 3. HTML directory XLSX files
    print("Processing HTML directory XLSX...")
    count = 0
    for xlsx in SOURCES['html_xlsx'].rglob("*.xlsx"):
        if 'scratch' in str(xlsx).lower() or xlsx.stem.startswith('~'):
            continue
        try:
            texts = parse_xlsx_file(xlsx)
            for t in texts:
                if t['id'] not in seen_ids:
                    all_texts.append(t)
                    seen_ids.add(t['id'])
                    count += 1
        except Exception as e:
            print(f"  Error processing {xlsx.name}: {e}")
    print(f"  Found {count} texts")
    
    # 4. Word Documents XLSX
    print("Processing Word Documents XLSX...")
    count = 0
    for xlsx in SOURCES['word_xlsx'].rglob("*.xlsx"):
        if xlsx.stem.startswith('~'):
            continue
        try:
            texts = parse_xlsx_file(xlsx)
            for t in texts:
                t['subcategory'] = 'Yom Kippur'
                if t['id'] not in seen_ids:
                    all_texts.append(t)
                    seen_ids.add(t['id'])
                    count += 1
        except Exception as e:
            print(f"  Error processing {xlsx.name}: {e}")
    print(f"  Found {count} texts")
    
    # 5. Converted HTML
    print("Processing converted HTML...")
    count = 0
    for xlsx in SOURCES['converted_html'].rglob("*.xlsx"):
        try:
            texts = parse_xlsx_file(xlsx)
            for t in texts:
                if t['id'] not in seen_ids:
                    all_texts.append(t)
                    seen_ids.add(t['id'])
                    count += 1
        except Exception as e:
            print(f"  Error processing {xlsx.name}: {e}")
    print(f"  Found {count} texts")
    
    # 6. Converted XML (Rosh Pinna)
    print("Processing converted XML...")
    count = 0
    for xlsx in SOURCES['converted_xml'].rglob("*.xlsx"):
        try:
            texts = parse_xlsx_file(xlsx)
            for t in texts:
                if t['id'] not in seen_ids:
                    all_texts.append(t)
                    seen_ids.add(t['id'])
                    count += 1
        except Exception as e:
            print(f"  Error processing {xlsx.name}: {e}")
    print(f"  Found {count} texts")
    
    return all_texts


def build_catalog(texts):
    """Build a catalog organized by category."""
    catalog = defaultdict(lambda: defaultdict(list))
    
    for text in texts:
        cat = text['category']
        subcat = text.get('subcategory', 'General')
        
        catalog[cat][subcat].append({
            'id': text['id'],
            'title_en': text['title_en'],
            'title_he': text.get('title_he', ''),
            'has_intro': 'introduction' in text
        })
    
    # Convert to regular dict and sort
    result = {}
    for cat in sorted(catalog.keys()):
        result[cat] = {}
        for subcat in sorted(catalog[cat].keys()):
            result[cat][subcat] = sorted(catalog[cat][subcat], key=lambda x: x['title_en'].lower())
    
    return result


def main():
    """Build the site."""
    print("="*60)
    print("BUILDING KARAITE TEXTS WEBSITE")
    print("="*60)
    
    # Create directories
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    (DATA_DIR / "texts").mkdir(exist_ok=True)
    
    # Collect all texts
    print("\nCollecting texts...")
    texts = collect_all_texts()
    print(f"\nTotal texts collected: {len(texts)}")
    
    # Build catalog
    print("\nBuilding catalog...")
    catalog = build_catalog(texts)
    
    # Save catalog
    with open(DATA_DIR / "catalog.json", 'w', encoding='utf-8') as f:
        json.dump(catalog, f, ensure_ascii=False, indent=2)
    print(f"  Saved catalog.json")
    
    # Save individual text files
    print("\nSaving individual texts...")
    for text in texts:
        text_path = DATA_DIR / "texts" / f"{text['id']}.json"
        with open(text_path, 'w', encoding='utf-8') as f:
            json.dump(text, f, ensure_ascii=False, indent=2)
    print(f"  Saved {len(texts)} text files")
    
    # Print summary
    print("\n" + "="*60)
    print("CATALOG SUMMARY")
    print("="*60)
    for cat, subcats in catalog.items():
        total = sum(len(items) for items in subcats.values())
        print(f"\n{cat}: {total} texts")
        for subcat, items in subcats.items():
            print(f"  {subcat}: {len(items)}")
    
    print(f"\n{'='*60}")
    print(f"Total: {len(texts)} texts")
    print(f"Output: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
