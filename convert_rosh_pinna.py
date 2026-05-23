#!/usr/bin/env python3
"""
Convert Rosh Pinna XML files (Hebrew, English, Arabic) to standard XLSX format.
Aligns paragraphs using ¶ markers and preserves footnotes and formatting.

Formatting markers preserved:
  - <i>text</i>     -> _text_      (italics)
  - <b>text</b>     -> **text**    (bold)
  - <u>text</u>     -> __text__    (underline)
  - <sc>text</sc>   -> ^^text^^    (small caps)
  - <green>text</green> -> {{bible:text}}  (biblical quote - Hebrew)
  - <quran>text</quran> -> {{quran:text}}  (Quranic quote - Arabic)
  - <sup>text</sup> -> ^(text)     (superscript)
  - <url>text</url> -> [[text]]    (URL/link)
  - <footnote id="n">text</footnote> -> [n] in text, footnote stored separately
"""

import re
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

XML_DIR = Path("/Users/shawn/knowledge-chatbot/documents/new_karaite/data_karaites/XML")
OUTPUT_DIR = Path("/Users/shawn/karaite-texts/xlsx_from_xml")

# Standard columns
COLUMNS = [
    "File Name", "Pattern", "Hebrew Name", "English Name", 
    "In Place of", "Reciter", "Hebrew Line #", "Hebrew Text",
    "English Transliteration", "English Translation", "Comments",
    "Time Starting", "Time Ending", "Arabic Text"
]


def normalize_marker(marker):
    """Normalize paragraph markers (¶1.1 -> ¶1.01)."""
    if not marker:
        return None
    # Handle ¶X.Y -> ¶X.0Y for single digit Y
    return re.sub(r'¶(\d+)\.(\d)$', r'¶\1.0\2', marker)


def extract_footnotes_inline(text):
    """
    Extract inline footnotes from text.
    Returns (clean_text, footnotes_list)
    
    Footnotes appear as: text<footnote id="21">footnote content</footnote>more text
    We extract them and mark their position with [n].
    """
    footnotes = []
    
    def replace_fn(match):
        fn_id = match.group(1)
        fn_text = match.group(2).strip()
        # Preserve formatting in footnote text too
        fn_text = convert_formatting_tags(fn_text)
        fn_text = re.sub(r'<[^>]+>', '', fn_text)  # Remove any remaining tags
        fn_text = re.sub(r'\s+', ' ', fn_text)
        footnotes.append((fn_id, fn_text))
        return f'[{fn_id}]'
    
    clean = re.sub(r'<footnote id="(\d+)">(.+?)</footnote>', replace_fn, text, flags=re.DOTALL)
    return clean, footnotes


def convert_formatting_tags(text):
    """Convert XML formatting tags to plain-text markers."""
    if not text:
        return text
    
    # Italics: <i>text</i> -> _text_
    text = re.sub(r'<i>([^<]*)</i>', r'_\1_', text)
    
    # Bold: <b>text</b> -> **text**
    text = re.sub(r'<b>([^<]*)</b>', r'**\1**', text)
    
    # Underline: <u>text</u> -> __text__
    text = re.sub(r'<u>([^<]*)</u>', r'__\1__', text)
    
    # Small caps: <sc>text</sc> -> ^^text^^
    text = re.sub(r'<sc>([^<]*)</sc>', r'^^\1^^', text)
    
    # Superscript: <sup>text</sup> -> ^(text)
    text = re.sub(r'<sup>([^<]*)</sup>', r'^(\1)', text)
    
    # Biblical quotes (Hebrew): <green>text</green> -> {{bible:text}}
    text = re.sub(r'<green>([^<]*)</green>', r'{{bible:\1}}', text)
    
    # Quranic quotes (Arabic): <quran>text</quran> -> {{quran:text}}
    text = re.sub(r'<quran>([^<]*)</quran>', r'{{quran:\1}}', text)
    
    # URLs: <url>text</url> -> [[text]]
    text = re.sub(r'<url>([^<]*)</url>', r'[[\1]]', text)
    
    # Center: <center>text</center> -> {{center:text}}
    text = re.sub(r'<center>([^<]*)</center>', r'{{center:\1}}', text)
    
    return text


def clean_text(text):
    """Clean XML text while preserving formatting markers."""
    if not text:
        return "", []
    
    # Extract footnotes first (before other processing)
    text, footnotes = extract_footnotes_inline(text)
    
    # Convert formatting tags to plain-text markers
    text = convert_formatting_tags(text)
    
    # Remove any remaining XML tags (margin, etc.)
    text = re.sub(r'<[^>]+>', '', text)
    
    # Clean whitespace (but preserve single spaces)
    text = re.sub(r'\s+', ' ', text).strip()
    
    return text, footnotes


def parse_xml_file(filepath):
    """
    Parse an XML file and extract paragraphs by marker.
    Returns dict: {marker: {'text': str, 'type': str, 'footnotes': list}}
    """
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    paragraphs = {}
    
    # Extract all content blocks (p, h1, quotation)
    # Pattern matches: <tag...><margin...>marker</margin>content</tag>
    # Or: <tag...>content</tag> (no margin)
    
    blocks = re.findall(
        r'<(p|h1|quotation)[^>]*>(?:<margin side="right">([^<]+)</margin>)?(.+?)</\1>',
        content, re.DOTALL
    )
    
    current_marker = None
    
    for tag, marker, text in blocks:
        marker = normalize_marker(marker) if marker else None
        
        if marker:
            current_marker = marker
        
        # Use current marker or generate one
        key = current_marker or f"_orphan_{len(paragraphs)}"
        
        clean, footnotes = clean_text(text)
        
        # Add type marker for headers and quotations
        if tag == 'h1':
            clean = f"{{{{header:{clean}}}}}"
        elif tag == 'quotation':
            clean = f"{{{{quote:{clean}}}}}"
        
        if key in paragraphs:
            # Append to existing (for quotations that follow a paragraph)
            paragraphs[key]['text'] += '\n' + clean
            paragraphs[key]['footnotes'].extend(footnotes)
        else:
            paragraphs[key] = {
                'text': clean,
                'type': tag,
                'footnotes': footnotes
            }
    
    return paragraphs


def create_xlsx():
    """Create XLSX from the three XML files."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # Parse all three files
    print("Parsing Hebrew XML...")
    hebrew = parse_xml_file(XML_DIR / 'rosh-pinna-hebrew.xml')
    print(f"  Found {len(hebrew)} paragraphs")
    
    print("Parsing English XML...")
    english = parse_xml_file(XML_DIR / 'rosh-pinna-english.xml')
    print(f"  Found {len(english)} paragraphs")
    
    print("Parsing Arabic XML...")
    arabic = parse_xml_file(XML_DIR / 'rosh-pinna-arabic.xml')
    print(f"  Found {len(arabic)} paragraphs")
    
    # Get all markers and sort them
    all_markers = sorted(
        set(hebrew.keys()) | set(english.keys()) | set(arabic.keys()),
        key=lambda x: (
            # Sort by chapter.paragraph numerically
            tuple(map(lambda p: float(p) if p.replace('.','').isdigit() else 0, 
                     x.replace('¶','').split('.'))) if x.startswith('¶') else (999, x)
        )
    )
    
    print(f"\nTotal unique markers: {len(all_markers)}")
    
    # Create workbook
    wb = Workbook()
    
    # ==================== TEXT SHEET ====================
    ws = wb.active
    ws.title = "Text"
    ws.append(COLUMNS)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    line_num = 0
    for marker in all_markers:
        if marker.startswith('_orphan'):
            continue
            
        line_num += 1
        
        heb_data = hebrew.get(marker, {'text': '', 'footnotes': []})
        eng_data = english.get(marker, {'text': '', 'footnotes': []})
        arb_data = arabic.get(marker, {'text': '', 'footnotes': []})
        
        # Combine footnotes from all languages
        all_footnotes = []
        if eng_data['footnotes']:
            for fn_id, fn_text in eng_data['footnotes']:
                all_footnotes.append(f"[{fn_id}] {fn_text}")
        
        comments = ' | '.join(all_footnotes) if all_footnotes else ''
        
        ws.append([
            "Rosh Pinna",           # File Name
            "",                      # Pattern
            "ראש פנה",              # Hebrew Name
            "Rosh Pinna",           # English Name
            "",                      # In Place of
            "",                      # Reciter
            line_num,               # Hebrew Line #
            heb_data['text'],       # Hebrew Text
            "",                      # English Transliteration (not applicable)
            eng_data['text'],       # English Translation
            comments,               # Comments (footnotes)
            "",                      # Time Starting
            "",                      # Time Ending
            arb_data['text']        # Arabic Text (extra column)
        ])
    
    # Adjust column widths
    ws.column_dimensions['H'].width = 60  # Hebrew
    ws.column_dimensions['J'].width = 60  # English
    ws.column_dimensions['K'].width = 50  # Comments
    ws.column_dimensions['N'].width = 60  # Arabic
    
    # ==================== FOOTNOTES SHEET ====================
    ws_fn = wb.create_sheet("Footnotes")
    ws_fn.append(["Footnote ID", "Paragraph", "Footnote Text"])
    for cell in ws_fn[1]:
        cell.font = Font(bold=True)
    
    # Collect all footnotes with their paragraph markers
    for marker in all_markers:
        eng_data = english.get(marker, {'text': '', 'footnotes': []})
        for fn_id, fn_text in eng_data['footnotes']:
            ws_fn.append([fn_id, marker, fn_text])
    
    ws_fn.column_dimensions['A'].width = 12
    ws_fn.column_dimensions['B'].width = 15
    ws_fn.column_dimensions['C'].width = 80
    
    # ==================== METADATA SHEET ====================
    ws_meta = wb.create_sheet("Metadata")
    ws_meta.append(["Field", "Value"])
    ws_meta['A1'].font = Font(bold=True)
    ws_meta['B1'].font = Font(bold=True)
    
    ws_meta.append(["Title (English)", "Rosh Pinna"])
    ws_meta.append(["Title (Hebrew)", "ראש פנה"])
    ws_meta.append(["Title (Arabic)", "رأس الزاوية"])
    ws_meta.append(["Category", "Halakhah"])
    ws_meta.append(["Languages", "Hebrew, English, Arabic"])
    ws_meta.append(["Source Files", "rosh-pinna-hebrew.xml, rosh-pinna-english.xml, rosh-pinna-arabic.xml"])
    ws_meta.append(["Total Paragraphs", str(line_num)])
    ws_meta.append(["Total Footnotes", str(sum(len(english.get(m, {'footnotes': []})['footnotes']) for m in all_markers))])
    ws_meta.append(["", ""])
    ws_meta.append(["--- FORMATTING MARKERS ---", ""])
    ws_meta.append(["_text_", "Italics"])
    ws_meta.append(["**text**", "Bold"])
    ws_meta.append(["__text__", "Underline"])
    ws_meta.append(["^^text^^", "Small Caps"])
    ws_meta.append(["^(text)", "Superscript"])
    ws_meta.append(["{{bible:text}}", "Biblical Quote (Hebrew)"])
    ws_meta.append(["{{quran:text}}", "Quranic Quote (Arabic)"])
    ws_meta.append(["{{header:text}}", "Section Header"])
    ws_meta.append(["{{quote:text}}", "Block Quote"])
    ws_meta.append(["{{center:text}}", "Centered Text"])
    ws_meta.append(["[[text]]", "URL/Link"])
    ws_meta.append(["[n]", "Footnote Reference (see Footnotes sheet)"])
    
    ws_meta.column_dimensions['A'].width = 25
    ws_meta.column_dimensions['B'].width = 60
    
    # ==================== ALIGNMENT CHECK SHEET ====================
    ws_align = wb.create_sheet("Alignment")
    ws_align.append(["Marker", "Hebrew", "English", "Arabic", "Notes"])
    for cell in ws_align[1]:
        cell.font = Font(bold=True)
    
    for marker in all_markers[:50]:  # First 50 for checking
        has_heb = "✓" if marker in hebrew else "✗"
        has_eng = "✓" if marker in english else "✗"
        has_arb = "✓" if marker in arabic else "✗"
        
        notes = []
        if marker not in hebrew:
            notes.append("Missing Hebrew")
        if marker not in english:
            notes.append("Missing English")
        if marker not in arabic:
            notes.append("Missing Arabic")
        
        ws_align.append([marker, has_heb, has_eng, has_arb, ", ".join(notes)])
    
    # Save
    output_path = OUTPUT_DIR / "Rosh Pinna.xlsx"
    wb.save(output_path)
    print(f"\nCreated: {output_path}")
    
    # Print summary
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"Total paragraphs: {line_num}")
    print(f"Footnotes extracted: {sum(len(english.get(m, {'footnotes': []})['footnotes']) for m in all_markers)}")
    print(f"Sheets created: Text, Footnotes, Metadata, Alignment")


if __name__ == "__main__":
    create_xlsx()
