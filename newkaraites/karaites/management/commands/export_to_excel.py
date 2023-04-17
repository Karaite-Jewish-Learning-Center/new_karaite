import re
from bs4 import BeautifulSoup
from django.core.management.base import BaseCommand
from ...models import (KaraitesBookAsArray, TableOfContents)
from openpyxl import Workbook
from pathlib import Path
from .command_utils.argments import arguments
from .process_arguments import process_arguments
from collections import OrderedDict

file_name = Path().cwd() / 'exported/'


class Command(BaseCommand):
    """Export book from database to excel"""

    def add_arguments(self, parser):
        arguments(parser)

    def handle(self, *args, **options):
        """Export book from database to excel"""
        query = process_arguments(options)
        if not query:
            return

        row = 1
        for q in query:
            title = q.book_title_en

            wb = Workbook()
            ws1 = wb.active
            ws1.title = title[0:30]

            ws2 = wb.create_sheet('Table of Contents')
            ws2.title = 'Table of Contents'

            toc_row = 1
            ws2.column_dimensions['A'].width = 15
            ws2.column_dimensions['B'].width = 100
            ws2[f'A{toc_row}'] = 'Paragraph Start'
            ws2[f'B{toc_row}'] = 'Toc'
            toc_row = 2

            # load table of contents
            toc_dic = OrderedDict()
            for toc in TableOfContents.objects.filter(karaite_book__book_title_en=title).order_by('karaite_book',
                                                                                                  'start_paragraph'):
                toc_dic[toc.start_paragraph] = toc.subject

            ws3 = wb.create_sheet('Introduction')
            ws3.title = 'Introduction'

            ws1.column_dimensions['A'].width = 15
            ws1.column_dimensions['B'].width = 100
            ws1[f'A{row}'] = 'Paragraph'
            ws1[f'B{row}'] = 'Hebrew'
            ws1[f'C{row}'] = 'English'
            ws1[f'D{row}'] = 'footnote Number'
            ws1[f'E{row}'] = 'footnote Hebrew'
            ws1[f'F{row}'] = 'footnote English'

            row += 1
            paragraphs_number = 1
            for book in KaraitesBookAsArray.objects.filter(book__book_title_en=title).order_by('book',
                                                                                               'paragraph_number'):
                html_he = book.book_text[2]
                html_en = book.book_text[0]
                soup_he = BeautifulSoup(html_he, 'html.parser')
                soup_en = BeautifulSoup(html_en, 'html.parser')

                text_he = soup_he.get_text().strip()
                # keep only one empty line between paragraphs
                if text_he == ' ' or text_he == '':
                    continue

                # update TOC and fix new number of paragraph in TOC
                if book.paragraph_number in toc_dic:
                    ws2[f'A{toc_row}'] = paragraphs_number
                    ws2[f'B{toc_row}'] = toc_dic[book.paragraph_number][0]
                    toc_row += 2

                # paragraph footnotes
                foot_notes_he = []
                for foot_note in soup_he.find_all('span', {'class': "en-foot-note"}):
                    foot_notes_he.append(foot_note.get('data-tip').strip().replace('\n', ' ').replace('\xa0', ''))

                text_he = soup_he.get_text().replace('\xa0', ' ').replace('\n', ' ').replace('\r', '')

                lines = text_he.split('.')
                period = '.' if len(lines) > 1 else ''

                for line in lines:

                    line = line.replace('\n', ' ')

                    if line == ' ' or line == '':
                        continue

                    ws1[f'A{row}'] = paragraphs_number
                    ws1[f'B{row}'] = line + period

                    # add footnotes, some paragraphs have more than one footnote
                    for fn in foot_notes_he:

                        foot_note_number = re.search(r'[(\d+)]', fn).group()

                        if foot_note_number:

                            found_in_line = re.findall(r'[(\d+)]', line)

                            if foot_note_number in found_in_line:
                                if ws1[f'D{row}'].value is None:
                                    # first footnote in the paragraph and possibly the only one
                                    ws1[f'D{row}'] = fn
                                else:
                                    # more than one footnote in the same paragraph
                                    ws1[f'D{row}'].value += '\n' + fn

                    row += 1
                    paragraphs_number += 1

                # add an empty line between paragraphs
                ws1[f'A{row}'] = paragraphs_number
                ws1[f'B{row}'] = ' '
                row += 1
                paragraphs_number += 1

                # sys.stdout.write(
                #     f'\rProcessing book:{book.book.book_title_en}  paragraph:{book.paragraph_number}       ')
                # sys.stdout.flush()
            wb.save(file_name / f'{title}.xlsx')
