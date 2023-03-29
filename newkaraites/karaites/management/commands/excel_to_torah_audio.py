from datetime import datetime
from django.core.files import File
from django.core.management.base import BaseCommand
from ...models import (BookAsArrayAudio,
                       AudioBook)
from openpyxl import load_workbook
import os

FILE_NAME = f'{os.getcwd()}/audioProject/torah_audio.xlsx'
AUDIO_DIR = f'{os.getcwd()}/audioProject/Parashat_bereshit-aliyot-mp3_2023-03-19_1724/'
OUT_CONST = f'{os.getcwd()}/frontend/src/constants/torahPortions.ts'

MAIN_TEMPLATE = f"""torahPortions = {{
    '{{book}}': {{
        {{data}}
    }}
}}
"""

DATA_TEMPLATE = f"""{{
'{{torah_portion}}': [
           {{data}},
        ]
        }},"""


class Command(BaseCommand):
    """ Read execl import Torah audio"""
    help = 'Import Torah audio from excel.'

    @staticmethod
    def fix_audio_length(time):
        if time is None:
            return None
        if len(time) == 11:
            time = datetime.strptime(time, '%H:%M:%S:%f')
            return time.strftime('%H:%M:%S.%f')[:-3]

    def handle(self, *args, **options):
        book = 'Genesis'
        wb = load_workbook(FILE_NAME)
        ws = wb[book]
        row = 2

        while True:
            if ws[f'A{row}'].value is None:
                break
            file_name = ws[f'A{row}'].value
            chapter = ws[f'F{row}'].value
            verse = ws[f'G{row}'].value
            start_time = ws[f'J{row}'].value
            end_time = ws[f'K{row}'].value
            print(f'chapter:{chapter} verse:{verse} start_time:{start_time} end_time:{end_time}')

            try:
                audio_book = AudioBook.objects.get(audio_name=file_name)
            except AudioBook.DoesNotExist:
                audio_book = AudioBook()
                audio_book.audio_name = file_name
                audio_book.audio_file.save(file_name, File(open(f'{AUDIO_DIR}{file_name}', 'rb')))
                audio_book.save()
            row += 1

            book = BookAsArrayAudio.objects.get(book__book_title_en='Genesis', chapter=chapter, verse=verse)
            book.audio = audio_book
            if start_time is not None:
                book.start = self.fix_audio_length(start_time)
            book.end = self.fix_audio_length(end_time)
            book.save()

        data = 'export const torahPortions ={'
        data += f"'{book}':{{"
        row = 2
        while ws[f'C{row}'].value is not None:

            torah_portion = ws[f'B{row}'].value
            data += f"'{torah_portion}':\n[\n"
            while ws[f'B{row}'].value == torah_portion:

                start = [int(ws[f'F{row}'].value), int(ws[f'G{row}'].value)]
                number = ws[f'C{row}'].value.split('.')[0]

                while ws[f'C{row}'].value is not None and number == ws[f'C{row}'].value.split('.')[0]:
                    stop = [int(ws[f'F{row}'].value), int(ws[f'G{row}'].value)]
                    traditional = ws[f'C{row}'].value.split('.')[1].strip()
                    row += 1

                start_stop = start + stop + [traditional]
                data += f"{{{number}: {start_stop}}},\n"

            data += '],\n'
        data += '},}\n'
        open(OUT_CONST, 'wb').write(data.encode('utf-8'))

        print('Done')
        print('Please open file torahPortions.ts and reformat it on your ide.')
        print('Run command update_audio.py')