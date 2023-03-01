from django.core.management.base import BaseCommand
from ...models import (Organization,
                       AudioBook,
                       BookAsArrayAudio)
from openpyxl import load_workbook


class Command(BaseCommand):
    help = 'Populate audio book marker for a specific book'

    @staticmethod
    def format_time(time):
        """ Format time to hh:mm:ss:ms """

        def pad_with_zero(s):
            return f'{s:0>2}'

        time = time.split('.')
        if len(time) == 4:
            time = [time[0], time[1], time[2], time[3]]
        if len(time) == 3:
            time = [0, time[0], time[1], time[2]]
        if len(time) == 2:
            time = [0, 0, time[0], time[1]]
        elif len(time) == 1:
            time = [0, 0, 0, time[0]]

        time_list = [pad_with_zero(x) for x in time[0:3]]

        return format(f'{time_list[0]}:{time_list[1]}:{time_list[2]}.{time[3]}')

    def add_arguments(self, parser):
        parser.add_argument(
            '--debug',
            default=False,
            action='store_true',
            help=help,
        )

        parser.add_argument(
            '--xls_file',
            default='',
            help='The excel file to read from',
        )

    def handle(self, *args, **options):
        """ Populate audio bible books with start and stop times for each chapter and verser """

        if options['xls_file'] == '':
            print('Please specify a file name')
            return

        wb = load_workbook(options['xls_file'])
        ws = wb.active

        audio_file = ws['A2'].value
        audio_name = audio_file.replace('.mp3', '').replace('.wav', '').replace('.ogg', '')
        book_title = ws['B2'].value
        try:
            organization = Organization.objects.get(book_title_en=book_title)
        except Organization.DoesNotExist:
            print(f'Could not find organization for book title {book_title}')
            return

        #  audiobook table
        try:
            audio_book = AudioBook.objects.get(audio_name=audio_name)
        except AudioBook.DoesNotExist:
            audio_book = AudioBook()
            audio_book.audio_name = audio_name
            audio_book.audio_file = f'audio_book/{audio_file}'
            audio_book.save()

        row = 2
        while ws[f'C{row}'].value is not None:
            chapter = int(ws[f'C{row}'].value)
            verse = int(ws[f'D{row}'].value)
            start = ws[f'E{row}'].value
            end = ws[f'F{row}'].value

            print(f'Processing chapter:{chapter} verse:{verse}')
            book = BookAsArrayAudio.objects.get(book=organization, chapter=chapter, verse=verse)
            book.audio = audio_book
            if start not in [None, '']:
                book.start_ms = book.convert_time_to_seconds(self.format_time(start))
            book.end_ms = book.convert_time_to_seconds(self.format_time(end))
            book.save()

            row += 1
