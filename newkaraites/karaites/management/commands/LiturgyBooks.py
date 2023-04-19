from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from ...models import (Songs,
                       LiturgyBook,
                       LiturgyDetails)
from django.core.files import File
from pathlib import Path


path = Path() / 'data_karaites/HTML/Liturgy/Shabbat Morning Services/Qedushot and Piyyut Parasha/'


class Command(BaseCommand):

    @staticmethod
    def save_data(liturgy_details, songs, line, line_number):
        # save liturgyBook
        try:
            liturgy_book = LiturgyBook.objects.get(book=liturgy_details, song=songs, line_number=line_number)
        except LiturgyBook.DoesNotExist:
            liturgy_book = LiturgyBook()

        liturgy_book.book = liturgy_details
        liturgy_book.song = songs
        liturgy_book.book_text = line
        liturgy_book.line_number = line_number
        liturgy_book.save()

    def add_arguments(self, parser):
        parser.add_argument(
            '--xls_file',
            default='Kedushot and Piyyut Parasha.xlsx',
            action='store_true',
            help='Update LiturgyBooks table with with excel data file.',
        )

    def handle(self, *args, **options):
        """ Read excel import songs , book text and audio, markers """
        books = ['Atta Qadosh']  # , 'Essa Lamerahoq', 'El Mistatter', 'Adir Venora', 'Ehad Elohenu']
        file_name = path / options['xls_file']

        wb = load_workbook(file_name)

        for book in books:
            ws = wb[book]

            # song details
            hebrew_name = ws['C2'].value
            english_name = ws['D2'].value
            LiturgyDetails.objects.filter(hebrew_name=book).delete()
            liturgy_details = LiturgyDetails()
            liturgy_details.occasion = ws['B2'].value
            liturgy_details.hebrew_name = hebrew_name
            liturgy_details.english_name = english_name
            liturgy_details.intro = ''
            liturgy_details.save()

            song_file = ws['A2'].value

            try:
                songs = Songs.objects.get(song_title=english_name)
            except Songs.DoesNotExist:
                songs = Songs()
            song_file_name = path / song_file
            songs.song_title = english_name
            songs.song_file.save(song_file, File(open(song_file_name, 'rb')))
            songs.save()

            row = 2
            line_number = 1
            english_translation = []
            hebrew_text = []
            while True:
                # line number
                if ws[f'I{row}'].value is None:
                    break

                # [hebrew, transliteration, english audio_start, audio_end, reciter, censored, line_number, comments]
                hebrew_text.append([
                    ws[f'J{row}'].value,
                    ws[f'K{row}'].value,
                    '',
                    ws[f'O{row}'].value,
                    ws[f'P{row}'].value,
                    ws[f'H{row}'].value,
                    ws[f'G{row}'].value,
                    ws[f'I{row}'].value,
                    ws[f'M{row}'].value
                ])

                english_translation.append(['', '', ws[f'L{row}'].value, '', '', '', '', '', ''])

                # end of verse
                if ws[f'F{row}'].value == '<end verse>':
                    # save hebrew text
                    for hebrew in hebrew_text:
                        self.save_data(liturgy_details, songs, hebrew, line_number)
                        line_number += 1

                    # one empty line between verses
                    separator = ['', '', '', '', '', '', '', '', '']
                    self.save_data(liturgy_details, songs, separator, line_number)

                    # save english translation
                    line_number += 1
                    for english in english_translation:
                        self.save_data(liturgy_details, songs, english, line_number)
                        line_number += 1

                    # one empty line between verses
                    self.save_data(liturgy_details, songs, separator, line_number)
                    line_number += 1
                    english_translation = []
                    hebrew_text = []

                row += 1

