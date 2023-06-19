from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from ...models import (Songs,
                       FirstLevel,
                       Classification,
                       KaraitesBookAsArray,
                       KaraitesBookDetails)
from ...utils import (Stack,
                      convert_time_string)

from django.core.files import File
from pathlib import Path

path = Path() / 'data_karaites/HTML/Liturgy/Shabbat Morning Services/Qedushot and Piyyut Parasha/'


class Command(BaseCommand):

    @staticmethod
    def save_data(liturgy_details, songs, line, line_number):
        # save liturgyBook
        try:
            liturgy_book = KaraitesBookAsArray.objects.get(book=liturgy_details, song=songs, line_number=line_number)
        except KaraitesBookAsArray.DoesNotExist:
            liturgy_book = KaraitesBookAsArray()

        liturgy_book.book = liturgy_details
        liturgy_book.song = songs
        liturgy_book.book_text = line
        liturgy_book.line_number = line_number
        liturgy_book.better_book = True
        liturgy_book.save()

    @staticmethod
    def save_song(english_name, song_file):
        try:
            songs = Songs.objects.get(song_title=english_name)
        except Songs.DoesNotExist:
            songs = Songs()
        song_file_name = path / song_file
        songs.song_title = english_name
        songs.save()

        return songs

    def add_arguments(self, parser):
        parser.add_argument(
            '--xls_file',
            default='Kedushot and Piyyut Parasha.xlsx',
            action='store_true',
            help='Update LiturgyBooks table with with excel data file.',
        )

    def handle(self, *args, **options):
        """ Read excel import songs , book text and audio, markers """
        books = ['Atta Qadosh', 'Essa Lamerahoq', 'El Mistatter', 'Adir Venora', 'Ehad Elohenu']
        file_name = path / options['xls_file']

        wb = load_workbook(file_name)
        stack = Stack()

        for book in books:
            ws = wb[book]

            # song details
            hebrew_name = ws['C2'].value
            english_name = ws['D2'].value
            song_file = ws['A2'].value

            KaraitesBookDetails.objects.filter(book_title_en=english_name).delete()

            liturgy_details = KaraitesBookDetails()
            liturgy_details.first_level = FirstLevel.objects.get(first_level='Liturgy')
            liturgy_details.book_classification = Classification.objects.get(
                classification_name='Shabbat Morning Services')
            liturgy_details.occasion = ws['B2'].value
            liturgy_details.book_title_he = hebrew_name
            liturgy_details.book_title_en = english_name
            liturgy_details.better_book = True
            liturgy_details.better_intro = []
            liturgy_details.better_toc = []
            liturgy_details.language = 'he'
            liturgy_details.author = None
            liturgy_details.save()

            liturgy_details.songs.add(self.save_song(english_name, song_file))

            row = 2
            line_number = 0
            spreadsheet_line = 1
            english_translation = []
            # use some empty lines on top to better display the text on the grid
            filler = ['', '', '', '', '', '', '', '', 0, '', 0, 1, 0]
            # hebrew_text = [filler, filler, filler, filler]
            hebrew_text = []
            # audio_start
            stack.push(convert_time_string(ws[f'O{row}'].value))
            # print('audio_start: ', ws[f'O{row}'].value, ' audio_end: ', ws[f'P{row}'].value)
            # input('Press Enter to continue...')

            while True:
                # line number
                if ws[f'I{row}'].value is None:
                    break
                # maybe more than one file song per book
                if ws[f'A{row}'].value is not None:
                    song_file = ws[f'A{row}'].value
                    songs = self.save_song(english_name, song_file)

                audio_start = stack.pop()
                stack.push(convert_time_string(ws[f'P{row}'].value))
                print('audio_start: ', audio_start, ' audio_end: ', stack.peek())
                # [hebrew, transliteration, english audio_start, audio_end, reciter, censored, line_number, comments,filler]
                hebrew_text.append([
                    ws[f'J{row}'].value,  # hebrew
                    ws[f'K{row}'].value,  # transliteration
                    '',
                    audio_start,
                    stack.peek(),  # audio_end
                    songs.id,  # song_id
                    ws[f'H{row}'].value,  # reciter
                    ws[f'G{row}'].value,  # censored
                    ws[f'I{row}'].value,  # line_number
                    ws[f'M{row}'].value,  # comments
                    0,  # end of verse, section or subtext? No
                    0,  # filler
                    0  # song end
                ])

                english_translation.append(['', '', ws[f'L{row}'].value, '', '', songs.id, '', '', '', '', 0, 1, 0])

                # end of verse, section or subtext
                end = ws[f'F{row}'].value
                if end is not None and ws[f'F{row}'].value.find('end') >= 0:
                    # save hebrew text
                    hebrew_text[-1][10] = 1  # end of verse, subtext
                    if end.find('<end subtext>') >= 0:
                        hebrew_text[-1][12] = 1  # end of song

                    for hebrew in hebrew_text:
                        self.save_data(liturgy_details, songs, hebrew, line_number)
                        line_number += 1

                    # save english translation
                    english_translation[-1][10] = 1  # end of verse, section or subtext? Yes

                    for english in english_translation:
                        self.save_data(liturgy_details, songs, english, line_number)
                        line_number += 1

                    english_translation = []
                    hebrew_text = []

                print('Processing  book: ', book, ' song: ', english_name, ' line_number: ', spreadsheet_line, )
                spreadsheet_line += 1
                row += 1

            # append some empty lines at bottom to better display the text on the grid
            for _ in range(10):
                self.save_data(liturgy_details, songs, filler, line_number)
                line_number += 1
