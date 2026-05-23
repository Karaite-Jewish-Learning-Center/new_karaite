#!/usr/bin/env python3
"""
Convert HTML files to standard XLSX format.
Groups related files (Text, Introduction, TOC) into single XLSX with multiple sheets.
"""

import re
import json
from pathlib import Path
from collections import defaultdict
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Paths
HTML_DIR = Path("/Users/shawn/knowledge-chatbot/documents/new_karaite/data_karaites/HTML")
OUTPUT_DIR = Path("/Users/shawn/karaite-texts/xlsx_from_html")

# Standard columns matching existing XLSX format
COLUMNS = [
    "File Name", "Pattern", "Hebrew Name", "English Name", 
    "In Place of", "Reciter", "Hebrew Line #", "Hebrew Text",
    "English Transliteration", "English Translation", "Comments",
    "Time Starting", "Time Ending"
]

def clean_text(text):
    """Clean whitespace from text."""
    if not text:
        return ""
    return re.sub(r'\s+', ' ', str(text)).strip()

def has_hebrew(text):
    """Check if text contains Hebrew characters."""
    return bool(re.search(r'[\u0590-\u05FF]', str(text))) if text else False

def extract_footnotes(soup):
    """Extract footnotes from HTML (Word-style _ftn1, _ftn2, etc.)."""
    footnotes = {}
    
    for a in soup.find_all('a'):
        name = a.get('name', '')
        if name and name.startswith('_ftn'):
            ftn_num = name.replace('_ftn', '')
            parent = a.find_parent(['p', 'div'])
            if parent:
                text = clean_text(parent.get_text())
                text = re.sub(r'^\s*\[\d+\]\s*', '', text)
                if text:
                    footnotes[ftn_num] = text
    
    return footnotes

def find_footnote_refs(text):
    """Find footnote markers like [1], [2] in text."""
    return re.findall(r'\[(\d+)\]', str(text))

def extract_table_content(soup, footnotes):
    """Extract content from HTML tables."""
    content = []
    
    for table in soup.find_all('table'):
        rows = table.find_all('tr')
        
        # First pass: determine table structure by examining first few rows
        # Count how many cells per row and identify column roles
        cell_counts = []
        for tr in rows[:10]:
            cells = tr.find_all('td')
            if cells and not cells[0].get('colspan'):
                cell_counts.append(len(cells))
        
        # Most common cell count
        typical_cols = max(set(cell_counts), key=cell_counts.count) if cell_counts else 2
        
        i = 0
        while i < len(rows):
            tr = rows[i]
            cells = tr.find_all('td')
            
            # Skip colspan rows (we'll grab English with Hebrew)
            if len(cells) == 1 or (cells and cells[0].get('colspan')):
                i += 1
                continue
            
            if len(cells) >= 2:
                hebrew = ""
                english = ""
                
                # Extract cell texts
                cell_texts = []
                for cell in cells:
                    paragraphs = cell.find_all('p')
                    if paragraphs:
                        texts = [clean_text(p.get_text()) for p in paragraphs if clean_text(p.get_text())]
                        cell_text = ' / '.join(texts)
                    else:
                        cell_text = clean_text(cell.get_text())
                    cell_texts.append(cell_text)
                
                # Handle based on number of columns
                if len(cells) >= 3:
                    # 3-column layout: Hebrew | English | Section Marker (ignore 3rd col)
                    # Or: Hebrew | Transliteration | English
                    for idx, cell_text in enumerate(cell_texts[:3]):
                        if not cell_text:
                            continue
                        if has_hebrew(cell_text):
                            hebrew = cell_text
                        elif idx == 1:
                            # Second column - could be English or transliteration
                            # If it looks like English prose, treat as English
                            english = cell_text
                        # Ignore 3rd column (usually section markers like #A, #1a)
                else:
                    # 2-column layout: Hebrew | English/Transliteration
                    for cell_text in cell_texts:
                        if not cell_text:
                            continue
                        if has_hebrew(cell_text):
                            hebrew = cell_text
                        else:
                            english = cell_text
                    
                    # Look for English in next row (for old 2-col + English-below layout)
                    if hebrew and not english and i + 1 < len(rows):
                        next_tr = rows[i + 1]
                        next_cells = next_tr.find_all('td')
                        if next_cells and (next_cells[0].get('colspan') or len(next_cells) == 1):
                            paragraphs = next_tr.find_all('p')
                            if paragraphs:
                                texts = [clean_text(p.get_text()) for p in paragraphs if clean_text(p.get_text())]
                                english = ' / '.join(texts)
                            else:
                                english = clean_text(next_tr.get_text())
                            i += 1
                
                if hebrew or english:
                    # Build comments from footnotes
                    all_text = f"{hebrew} {english}"
                    ftn_refs = find_footnote_refs(all_text)
                    comments = []
                    for ref in ftn_refs:
                        if ref in footnotes:
                            comments.append(f"[{ref}] {footnotes[ref]}")
                    
                    content.append({
                        "hebrew": hebrew,
                        "transliteration": "",  # HTML files don't have transliteration
                        "english": english,
                        "comments": " | ".join(comments) if comments else ""
                    })
            
            i += 1
    
    return content

def extract_prose_content(soup):
    """Extract prose content (for Introduction files)."""
    # Remove script/style
    for tag in soup(['script', 'style', 'head']):
        tag.decompose()
    
    # Get all paragraphs
    paragraphs = []
    for p in soup.find_all(['p', 'div']):
        text = clean_text(p.get_text())
        if text and len(text) > 10:  # Skip very short snippets
            paragraphs.append(text)
    
    return '\n\n'.join(paragraphs)

def extract_toc_content(soup):
    """Extract table of contents."""
    toc = []
    
    # Look for tables first
    for table in soup.find_all('table'):
        for row in table.find_all('tr'):
            cells = row.find_all('td')
            if len(cells) >= 2:
                hebrew = ""
                english = ""
                for cell in cells:
                    text = clean_text(cell.get_text())
                    if has_hebrew(text):
                        hebrew = text
                    elif text:
                        english = text
                if hebrew or english:
                    toc.append({"hebrew": hebrew, "english": english})
    
    # If no table, look for list items or paragraphs
    if not toc:
        for elem in soup.find_all(['li', 'p']):
            text = clean_text(elem.get_text())
            if text:
                if has_hebrew(text):
                    toc.append({"hebrew": text, "english": ""})
                else:
                    toc.append({"hebrew": "", "english": text})
    
    return toc

def extract_titles(soup):
    """Extract Hebrew and English titles."""
    title_en = ""
    title_he = ""
    
    for p in soup.find_all('p')[:15]:
        style = str(p.get('style', '')) or ''
        align = p.get('align', '')
        if 'center' in style or align == 'center':
            text = clean_text(p.get_text())
            if has_hebrew(text) and not title_he:
                title_he = text
            elif text and not title_en and len(text) < 100 and not has_hebrew(text):
                title_en = text
    
    return title_en, title_he

def group_related_files(html_files):
    """Group related HTML files by base name (Text, Introduction, TOC)."""
    groups = defaultdict(dict)
    
    for html_file in html_files:
        name = html_file.stem
        
        # Determine type and base name
        if '-Introduction' in name or 'Introduction' in name or 'Intro' in name:
            base = re.sub(r'[-_]?(Introduction|Intro)$', '', name, flags=re.IGNORECASE)
            groups[base]['intro'] = html_file
        elif '-TOC' in name or 'TOC' in name:
            base = re.sub(r'[-_]?TOC$', '', name, flags=re.IGNORECASE)
            groups[base]['toc'] = html_file
        elif '-Hebrew-English' in name:
            base = name.replace('-Hebrew-English', '')
            groups[base]['text'] = html_file
        elif '-Hebrew' in name:
            base = name.replace('-Hebrew', '')
            groups[base]['text'] = html_file
        else:
            # Standalone file
            groups[name]['text'] = html_file
    
    return groups

def create_xlsx_for_book(base_name, files, category, output_path):
    """Create a single XLSX with multiple sheets for a book."""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    title_en = base_name
    title_he = ""
    has_content = False
    
    # Process Text file
    if 'text' in files:
        with open(files['text'], 'r', encoding='utf-8', errors='ignore') as f:
            soup = BeautifulSoup(f.read(), 'html.parser')
        
        t_en, t_he = extract_titles(soup)
        if t_en:
            title_en = t_en
        if t_he:
            title_he = t_he
        
        footnotes = extract_footnotes(soup)
        content = extract_table_content(soup, footnotes)
        
        if content:
            has_content = True
            ws = wb.create_sheet("Text")
            ws.append(COLUMNS)
            
            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True)
            
            for i, row in enumerate(content, 1):
                ws.append([
                    base_name,      # File Name
                    "",             # Pattern
                    title_he,       # Hebrew Name
                    title_en,       # English Name
                    "",             # In Place of
                    "",             # Reciter
                    i,              # Hebrew Line #
                    row["hebrew"],                    # Column H - Hebrew Text
                    row["transliteration"],           # Column I - English Transliteration
                    row["english"],                   # Column J - English Translation
                    row["comments"],                  # Column K - Comments
                    "",             # Time Starting
                    ""              # Time Ending
                ])
            
            # Adjust column widths
            ws.column_dimensions['H'].width = 50
            ws.column_dimensions['I'].width = 50
            ws.column_dimensions['J'].width = 60
            ws.column_dimensions['K'].width = 40
    
    # Process Introduction file
    if 'intro' in files:
        with open(files['intro'], 'r', encoding='utf-8', errors='ignore') as f:
            soup = BeautifulSoup(f.read(), 'html.parser')
        
        intro_text = extract_prose_content(soup)
        
        if intro_text:
            has_content = True
            ws = wb.create_sheet("Introduction")
            ws.append(["Introduction Text"])
            ws['A1'].font = Font(bold=True)
            
            # Split into paragraphs and add rows
            for para in intro_text.split('\n\n'):
                if para.strip():
                    ws.append([para.strip()])
            
            ws.column_dimensions['A'].width = 100
    
    # Process TOC file
    if 'toc' in files:
        with open(files['toc'], 'r', encoding='utf-8', errors='ignore') as f:
            soup = BeautifulSoup(f.read(), 'html.parser')
        
        toc_content = extract_toc_content(soup)
        
        if toc_content:
            has_content = True
            ws = wb.create_sheet("TOC")
            ws.append(["Hebrew", "English", "Line Number"])
            for cell in ws[1]:
                cell.font = Font(bold=True)
            
            for item in toc_content:
                ws.append([item["hebrew"], item["english"], ""])
            
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 50
    
    # Add Metadata sheet
    ws = wb.create_sheet("Metadata")
    ws.append(["Field", "Value"])
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws.append(["Title (English)", title_en])
    ws.append(["Title (Hebrew)", title_he])
    ws.append(["Category", category])
    ws.append(["Source Files", ", ".join([f.name for f in files.values()])])
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60
    
    if has_content:
        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return True
    
    return False

def main():
    """Main processing function."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # Get all XLSX stems to skip files that already have XLSX
    xlsx_stems = set()
    
    # From kedushot.xlsx
    from openpyxl import load_workbook
    kedushot_path = Path("/Users/shawn/knowledge-chatbot/documents/new_karaite/karaites/management/commands/kedushot.xlsx")
    if kedushot_path.exists():
        wb = load_workbook(kedushot_path, read_only=True)
        for sheet in wb.sheetnames:
            if sheet not in ['Index Format', 'Info', 'DND']:
                xlsx_stems.add(re.sub(r'[^a-z0-9]', '', sheet.lower()))
        wb.close()
    
    # From data_karaites
    base = Path("/Users/shawn/knowledge-chatbot/documents/new_karaite/data_karaites")
    for xlsx in base.rglob("*.xlsx"):
        if 'scratch' not in str(xlsx).lower() and '_old' not in xlsx.stem.lower():
            xlsx_stems.add(re.sub(r'[^a-z0-9]', '', xlsx.stem.lower()))
    
    print(f"Found {len(xlsx_stems)} existing XLSX sources\n")
    
    # Process HTML files by category
    converted = 0
    skipped = 0
    
    for category_dir in HTML_DIR.iterdir():
        if not category_dir.is_dir():
            continue
        if 'scratch' in category_dir.name.lower():
            continue
        
        category = category_dir.name
        print(f"\n{'='*60}")
        print(f"Category: {category}")
        print('='*60)
        
        # Find all HTML files in this category
        html_files = list(category_dir.rglob("*.html"))
        
        # Filter out files that already have XLSX
        html_to_process = []
        for html in html_files:
            stem_norm = re.sub(r'[^a-z0-9]', '', html.stem.lower())
            # Remove intro/toc/hebrew suffixes for matching
            stem_base = re.sub(r'(introduction|intro|toc|hebrew|english|hebrewenglish)$', '', stem_norm)
            
            has_xlsx = any(stem_base in x or x in stem_base for x in xlsx_stems)
            if not has_xlsx:
                html_to_process.append(html)
        
        if not html_to_process:
            print(f"  All files have XLSX, skipping")
            continue
        
        # Group related files
        groups = group_related_files(html_to_process)
        
        for base_name, files in groups.items():
            print(f"\n  {base_name}:")
            for ftype, fpath in files.items():
                print(f"    - {ftype}: {fpath.name}")
            
            # Create output path
            output_path = OUTPUT_DIR / category / f"{base_name}.xlsx"
            
            if create_xlsx_for_book(base_name, files, category, output_path):
                print(f"    → Created: {output_path.name}")
                converted += 1
            else:
                print(f"    → No content extracted")
                skipped += 1
    
    print(f"\n{'='*60}")
    print(f"Done!")
    print(f"  Converted: {converted}")
    print(f"  Skipped (no content): {skipped}")
    print(f"  Output: {OUTPUT_DIR}")

if __name__ == "__main__":
    main()
