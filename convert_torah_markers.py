"""Merge Torah audio markers from the project xlsx into the Tanakh JSON files.

Reads `Torah Audio Recording Markers.xlsx`, copies the MP3s from
`Combined Parashiot by Aliya Folder/` into `site/audio/torah/` with slugified
names, and writes per-verse `audio` + `timing` and per-chapter `audioSegments`
into `site/data/tanakh/<book>.json`.

Sheets without a populated `File name`/audio column are skipped silently so
this script can be re-run as more sheets get filled in.
"""

from __future__ import annotations

import json
import re
import shutil
from collections import OrderedDict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX_PATH = ROOT / "Updated Torah Audio Recordings Markers.xlsx"
SOURCE_AUDIO_DIR = ROOT / "Combined Parashiot by Aliya Folder"
SITE_AUDIO_DIR = ROOT / "site" / "audio" / "torah"
TANAKH_DIR = ROOT / "site" / "data" / "tanakh"

FRAME_RATE = 30.0  # SMPTE non-drop assumed for HH:MM:SS:FF markers

BOOK_TO_FILE = {
    "Genesis": "genesis.json",
    "Exodus": "exodus.json",
    "Leviticus": "leviticus.json",
    "Numbers": "numbers.json",
    "Deuteronomy": "deuteronomy.json",
}


def smpte_to_seconds(value: str | None) -> float | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    parts = s.split(":")
    if len(parts) != 4:
        return None
    try:
        hh, mm, ss, ff = (int(p) for p in parts)
    except ValueError:
        return None
    return hh * 3600 + mm * 60 + ss + ff / FRAME_RATE


def _strip_mp3(name: str) -> str:
    return re.sub(r"\.mp3$", "", name, flags=re.IGNORECASE)


_ALIYAH_RE = re.compile(r"^\s*(\d+)\b.*?\b(\d+)[\s.]*([A-Za-z'_]+)", re.UNICODE)


def normalize_key(name: str) -> str:
    """Canonical key used to bridge xlsx filenames and disk filenames.

    The xlsx and disk disagree on parasha spelling (Noah vs Noach,
    Lech vs Lehc, Yisre'eli vs Yisre_eli) and on whether '.mp3' is
    included, so we key on `<parasha-number>-<aliyah-number>` (e.g. '01-3').
    """
    stem = _strip_mp3(name)
    m = _ALIYAH_RE.match(stem)
    if m:
        parasha_num = int(m.group(1))
        aliyah_num = int(m.group(2))
        return f"{parasha_num:02d}-{aliyah_num}"
    stem = stem.replace("'", "").replace("_", "")
    stem = re.sub(r"[^A-Za-z0-9]+", " ", stem).strip().lower()
    return re.sub(r"\s+", " ", stem)


def slugify_audio_name(disk_filename: str) -> str:
    """Turn the on-disk MP3 name into a URL-safe slug under site/audio/torah/.

    Examples:
        '01 Parashat Bereshit 1. Cohen.mp3'  -> '01-bereshit-1-cohen.mp3'
        '07 Parashat Vayyetze 7. Yisre_eli.mp3' -> '07-vayyetze-7-yisreeli.mp3'
        '13 Shemot 6. Yisre_eli.mp3'         -> '13-shemot-6-yisreeli.mp3'
        '03 Parashat Lehc Lecha 1. Cohen.mp3' -> '03-lech-lecha-1-cohen.mp3'
    """
    stem = _strip_mp3(disk_filename)
    stem = re.sub(r"\bParashat\b", "", stem, flags=re.IGNORECASE)
    stem = stem.replace("_", "")
    stem = re.sub(r"\bLehc\b", "Lech", stem)
    stem = re.sub(r"[^A-Za-z0-9]+", "-", stem)
    stem = re.sub(r"-+", "-", stem).strip("-").lower()
    return f"{stem}.mp3"


_disk_index: dict[str, str] | None = None


def _get_disk_index() -> dict[str, str]:
    global _disk_index
    if _disk_index is None:
        idx: dict[str, str] = {}
        for p in SOURCE_AUDIO_DIR.iterdir():
            if p.is_file() and p.suffix.lower() == ".mp3":
                idx[normalize_key(p.name)] = p.name
        _disk_index = idx
    return _disk_index


def copy_audio(xlsx_filename: str) -> str | None:
    """Copy the MP3 into site/audio/torah/ with a slugified name; return slug.

    Returns None if no matching MP3 exists on disk (recording not yet provided).
    """
    disk_index = _get_disk_index()
    key = normalize_key(xlsx_filename)
    disk_name = disk_index.get(key)
    if disk_name is None:
        return None
    src = SOURCE_AUDIO_DIR / disk_name
    SITE_AUDIO_DIR.mkdir(parents=True, exist_ok=True)
    slug = slugify_audio_name(disk_name)
    dst = SITE_AUDIO_DIR / slug
    if not dst.exists() or dst.stat().st_size != src.stat().st_size:
        shutil.copy2(src, dst)
    return slug


def parse_sheet(ws) -> list[dict]:
    """Return a list of marker rows for sheets that include a File name column.

    Returns an empty list for sheets that only have text (no audio yet).
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    if "file name" not in header:
        return []

    idx = {name: header.index(name) for name in (
        "file name",
        "torah portion",
        "traditional shabbat aliyah",
        "chapter",
        "verse",
        "time starts",
        "time ends",
    ) if name in header}

    markers: list[dict] = []
    last_end_by_file: dict[str, float | None] = {}

    for row in rows[1:]:
        filename = row[idx["file name"]]
        chapter = row[idx["chapter"]]
        verse = row[idx["verse"]]
        if filename is None or chapter is None or verse is None:
            continue
        filename = str(filename).strip()
        try:
            chapter = int(chapter)
            verse = int(verse)
        except (TypeError, ValueError):
            continue

        start = smpte_to_seconds(row[idx["time starts"]])
        end = smpte_to_seconds(row[idx["time ends"]])
        if start is None:
            start = last_end_by_file.get(filename, 0.0) or 0.0
        if end is None:
            continue
        last_end_by_file[filename] = end

        markers.append({
            "filename": filename,
            "parasha": (str(row[idx["torah portion"]]).strip()
                        if "torah portion" in idx and row[idx["torah portion"]] else None),
            "aliyah": (str(row[idx["traditional shabbat aliyah"]]).strip()
                       if "traditional shabbat aliyah" in idx and row[idx["traditional shabbat aliyah"]] else None),
            "chapter": chapter,
            "verse": verse,
            "start": round(start, 3),
            "end": round(end, 3),
        })

    return markers


def merge_into_book(book_path: Path, markers: list[dict]) -> tuple[int, int]:
    if not markers:
        return 0, 0

    with book_path.open("r", encoding="utf-8") as f:
        book = json.load(f)

    file_to_slug: dict[str, str] = {}

    by_chapter: dict[int, list[dict]] = {}
    for m in markers:
        by_chapter.setdefault(m["chapter"], []).append(m)

    verses_updated = 0
    chapters_with_audio: set[int] = set()

    for chapter_obj in book.get("chapters", []):
        cnum = chapter_obj.get("chapter")
        chapter_markers = by_chapter.get(cnum, [])
        if not chapter_markers:
            chapter_obj.pop("audioSegments", None)
            for v in chapter_obj.get("verses", []):
                for k in ("audio", "timing", "aliyah", "parasha"):
                    v.pop(k, None)
            continue

        marker_by_verse = {m["verse"]: m for m in chapter_markers}

        segments: "OrderedDict[str, dict]" = OrderedDict()
        missing_files: set[str] = set()
        for m in chapter_markers:
            if m["filename"] in file_to_slug:
                slug = file_to_slug[m["filename"]]
            else:
                slug = copy_audio(m["filename"])
                file_to_slug[m["filename"]] = slug or ""
            if not slug:
                missing_files.add(m["filename"])
                continue
            url = f"audio/torah/{slug}"
            label_parts = [m["parasha"] or "", m["aliyah"] or ""]
            label = " · ".join(p for p in label_parts if p) or slug
            if url not in segments:
                segments[url] = {"label": label, "url": url}

        for v in chapter_obj.get("verses", []):
            m = marker_by_verse.get(v.get("verse"))
            if not m or not file_to_slug.get(m["filename"]):
                for k in ("audio", "timing", "aliyah", "parasha"):
                    v.pop(k, None)
                continue
            slug = file_to_slug[m["filename"]]
            v["audio"] = f"audio/torah/{slug}"
            v["timing"] = {"start": m["start"], "end": m["end"]}
            if m["aliyah"]:
                v["aliyah"] = m["aliyah"]
            if m["parasha"]:
                v["parasha"] = m["parasha"]
            verses_updated += 1

        if segments:
            chapter_obj["audioSegments"] = list(segments.values())
            chapters_with_audio.add(cnum)
        else:
            chapter_obj.pop("audioSegments", None)
        for missing in missing_files:
            print(f"    [warn] no MP3 on disk for xlsx entry {missing!r} (skipped)")

    with book_path.open("w", encoding="utf-8") as f:
        json.dump(book, f, ensure_ascii=False, indent=2)

    return verses_updated, len(chapters_with_audio)


def main() -> None:
    if not XLSX_PATH.exists():
        raise SystemExit(f"Missing xlsx: {XLSX_PATH}")
    if not SOURCE_AUDIO_DIR.exists():
        raise SystemExit(f"Missing audio source dir: {SOURCE_AUDIO_DIR}")

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        if sheet_name not in BOOK_TO_FILE:
            print(f"  skip sheet {sheet_name!r} (not a Torah book)")
            continue
        book_path = TANAKH_DIR / BOOK_TO_FILE[sheet_name]
        if not book_path.exists():
            print(f"  skip {sheet_name}: missing {book_path}")
            continue
        markers = parse_sheet(wb[sheet_name])
        if not markers:
            print(f"  {sheet_name}: no audio markers in sheet (skipping)")
            continue
        verses, chapters = merge_into_book(book_path, markers)
        print(f"  {sheet_name}: {verses} verses, {chapters} chapters with audio")


if __name__ == "__main__":
    main()
