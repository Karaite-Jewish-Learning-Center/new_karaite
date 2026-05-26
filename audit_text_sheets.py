"""Audit Exodus / Leviticus / Numbers / Deuteronomy sheets for text-data issues.

These sheets currently have no audio markers, but they do have Chapter / Verse /
English / Hebrew columns. This checks for missing or duplicate verses, gaps,
out-of-range chapters or verses, and empty Hebrew/English cells.
"""

from __future__ import annotations
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX_PATH = ROOT / "Updated Torah Audio Recordings Markers.xlsx"

# Standard Masoretic chapter:verse counts.
COUNTS = {
    "Exodus": {1:22,2:25,3:22,4:31,5:23,6:30,7:29,8:28,9:35,10:29,11:10,12:51,
               13:22,14:31,15:27,16:36,17:16,18:27,19:25,20:23,21:37,22:30,
               23:33,24:18,25:40,26:37,27:21,28:43,29:46,30:38,31:18,32:35,
               33:23,34:35,35:35,36:38,37:29,38:31,39:43,40:38},
    "Leviticus": {1:17,2:16,3:17,4:35,5:26,6:23,7:38,8:36,9:24,10:20,11:47,
                  12:8,13:59,14:57,15:33,16:34,17:16,18:30,19:37,20:27,
                  21:24,22:33,23:44,24:23,25:55,26:46,27:34},
    "Numbers": {1:54,2:34,3:51,4:49,5:31,6:27,7:89,8:26,9:23,10:36,11:35,
                12:16,13:33,14:45,15:41,16:35,17:28,18:32,19:22,20:29,
                21:35,22:41,23:30,24:25,25:18,26:65,27:23,28:31,29:40,
                30:17,31:54,32:42,33:56,34:29,35:34,36:13},
    "Deuteronomy": {1:46,2:37,3:29,4:49,5:30,6:25,7:26,8:20,9:29,10:22,
                    11:32,12:31,13:19,14:29,15:23,16:22,17:20,18:22,19:21,
                    20:20,21:23,22:29,23:26,24:22,25:19,26:19,27:26,28:69,
                    29:28,30:20,31:30,32:52,33:29,34:12},
}


def audit(sheet_name: str, ws, expected: dict[int, int]):
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    if "chapter" not in header or "verse" not in header:
        return None
    ci = header.index("chapter"); vi = header.index("verse")
    ei = header.index("english") if "english" in header else None
    hi = header.index("hebrew") if "hebrew" in header else None

    seen: dict[tuple[int, int], int] = {}
    row_issues: list[str] = []

    def get(row, idx):
        return row[idx] if idx is not None and idx < len(row) else None

    for line_no, row in enumerate(rows[1:], start=2):
        ch = get(row, ci)
        v = get(row, vi)
        if ch is None and v is None:
            continue
        if ch is None or v is None:
            row_issues.append(f"row {line_no}: partial row ch={ch!r} v={v!r}")
            continue
        try:
            ch = int(ch); v = int(v)
        except (TypeError, ValueError):
            row_issues.append(f"row {line_no}: non-numeric ch={row[ci]!r} v={row[vi]!r}")
            continue

        key = (ch, v)
        if key in seen:
            row_issues.append(
                f"row {line_no}: duplicate {ch}:{v} (also row {seen[key]})"
            )
        seen[key] = line_no

        if ch < 1 or ch > max(expected):
            row_issues.append(
                f"row {line_no}: chapter {ch} out of range 1..{max(expected)}"
            )
            continue
        if v < 1 or v > expected.get(ch, 0):
            row_issues.append(
                f"row {line_no}: {ch}:{v} verse out of range (Masoretic max {expected.get(ch)})"
            )

        eng = get(row, ei)
        if ei is not None and (eng is None or (isinstance(eng, str) and not eng.strip())):
            row_issues.append(f"row {line_no}: {ch}:{v} missing English")
        heb = get(row, hi)
        if hi is not None and (heb is None or (isinstance(heb, str) and not heb.strip())):
            row_issues.append(f"row {line_no}: {ch}:{v} missing Hebrew")

    missing_per_chapter: list[tuple[int, list[int]]] = []
    extras: list[str] = []
    for ch, max_v in expected.items():
        present = {v for (c, v) in seen if c == ch}
        miss = sorted(set(range(1, max_v + 1)) - present)
        extra = sorted(present - set(range(1, max_v + 1)))
        if miss:
            missing_per_chapter.append((ch, miss))
        if extra:
            extras.append(f"chapter {ch}: unexpected verses {extra}")

    chs_present = sorted({c for (c, v) in seen})
    return {
        "rows_seen": len(seen),
        "chapters_present": (chs_present[0], chs_present[-1]) if chs_present else None,
        "row_issues": row_issues,
        "missing_per_chapter": missing_per_chapter,
        "extras": extras,
    }


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    for sheet_name, expected in COUNTS.items():
        if sheet_name not in wb.sheetnames:
            print(f"\n=== {sheet_name} === (sheet missing)")
            continue
        result = audit(sheet_name, wb[sheet_name], expected)
        print(f"\n=== {sheet_name} ===")
        if result is None:
            print("  could not audit (no chapter/verse columns)")
            continue
        rng = result["chapters_present"]
        print(f"  rows: {result['rows_seen']} verses, "
              f"chapters present: {rng[0]}..{rng[1] if rng else '-'} of 1..{max(expected)}")

        if not result["row_issues"] and not result["missing_per_chapter"] and not result["extras"]:
            print("  no issues")
            continue

        if result["row_issues"]:
            print(f"  row-level issues ({len(result['row_issues'])}):")
            for line in result["row_issues"][:60]:
                print(f"    {line}")
            if len(result["row_issues"]) > 60:
                print(f"    ... and {len(result['row_issues']) - 60} more")

        if result["missing_per_chapter"]:
            full = [(ch, miss) for ch, miss in result["missing_per_chapter"]
                    if len(miss) == expected[ch]]
            partial = [(ch, miss) for ch, miss in result["missing_per_chapter"]
                       if len(miss) != expected[ch]]
            if full:
                runs = []
                run_start = run_prev = None
                for ch, _ in full:
                    if run_start is None:
                        run_start = run_prev = ch
                    elif ch == run_prev + 1:
                        run_prev = ch
                    else:
                        runs.append((run_start, run_prev))
                        run_start = run_prev = ch
                runs.append((run_start, run_prev))
                ranges = ", ".join(
                    f"{a}-{b}" if a != b else str(a) for a, b in runs
                )
                print(f"  chapters entirely missing: {ranges}")
            if partial:
                print(f"  chapters with partial gaps:")
                for ch, miss in partial:
                    if len(miss) <= 12:
                        print(f"    Ch {ch}: missing {miss}")
                    else:
                        print(f"    Ch {ch}: missing {len(miss)} verses ({miss[:5]} ... {miss[-3:]})")

        if result["extras"]:
            print(f"  extras:")
            for line in result["extras"]:
                print(f"    {line}")


if __name__ == "__main__":
    main()
