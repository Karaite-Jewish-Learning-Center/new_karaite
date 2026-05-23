#!/usr/bin/env python3
"""
Convert prose-based HTML files (no tables) to XLSX format.
Handles:
1. Verse-marked texts (like Deuteronomy commentary) - aligns Hebrew/English by verse markers
2. Prose texts (like Adderet Eliyahu) - sequential paragraphs

Formatting markers preserved:
  - _text_       Italics
  - **text**     Bold  
  - __text__     Underline
  - {{bible:}}   Biblical quote
  - {{header:}}  Section header
"""

import re
from pathlib import Path
from collections import defaultdict
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font

HTML_DIR = Path("/Users/shawn/knowledge-chatbot/documents/new_karaite/data_karaites/HTML")
OUTPUT_DIR = Path("/Users/shawn/karaite-texts/xlsx_from_html")

COLUMNS = [
    "File Name", "Pattern", "Hebrew Name", "English Name", 
    "In Place of", "Reciter", "Hebrew Line #", "Hebrew Text",
    "English Transliteration", "English Translation", "Comments",
    "Time Starting", "Time Ending"
]


def has_hebrew(text):
    """Check if text contains Hebrew characters."""
    return bool(re.search(r'[\u0590-\u05FF]', str(text))) if text else False


def convert_formatting(text):
    """Convert HTML formatting tags to plain-text markers."""
    if not text:
        return ""
    
    # Bold
    text = re.sub(r'<b>([^<]*)</b>', r'**\1**', text)
    text = re.sub(r'<strong>([^<]*)</strong>', r'**\1**', text)
    
    # Italics
    text = re.sub(r'<i>([^<]*)</i>', r'_\1_', text)
    text = re.sub(r'<em>([^<]*)</em>', r'_\1_', text)
    
    # Underline
    text = re.sub(r'<u>([^<]*)</u>', r'__\1__', text)
    
    # Superscript (for footnote markers)
    text = re.sub(r'<sup>([^<]*)</sup>', r'^(\1)', text)
    
    return text


def clean_text(text):
    """Clean HTML text while preserving formatting markers."""
    if not text:
        return ""
    
    # Convert formatting first
    text = convert_formatting(text)
    
    # Remove remaining HTML tags
    text = re.sub(r'<[^>]+>', '', text)
    
    # Clean whitespace
    text = re.sub(r'\s+', ' ', text).strip()
    
    return text


def extract_verse_marker(text):
    """Extract verse marker like '1:1' or '1:2-3' from start of text."""
    match = re.match(r'^(\d+:\d+(?:-\d+)?)\s*', text)
    if match:
        return match.group(1), text[match.end():].strip()
    return None, text


def parse_verse_marked_html(filepath):
    """
    Parse HTML with verse markers (e.g., Deuteronomy commentary).
    Returns dict: {verse_marker: text}
    """
    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()
    
    soup = BeautifulSoup(content, 'html.parser')
    
    verses = {}
    current_verse = None
    current_text = []
    
    for p in soup.find_all('p'):
        text = clean_text(p.get_text())
        if not text:
            continue
        
        verse, remainder = extract_verse_marker(text)
        
        if verse:
            # Save previous verse
            if current_verse and current_text:
                verses[current_verse] = ' '.join(current_text)
            
            current_verse = verse
            current_text = [remainder] if remainder else []
        elif current_verse:
            # Continue current verse
            current_text.append(text)
        else:
            # Text before first verse marker - use line number
            if text.strip():
                verses[f"_intro_{len(verses)}"] = text
    
    # Save last verse
    if current_verse and current_text:
        verses[current_verse] = ' '.join(current_text)
    
    return verses


def parse_prose_html(filepath):
    """
    Parse prose HTML (no verse markers).
    Returns list of paragraphs with metadata.
    """
    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()
    
    soup = BeautifulSoup(content, 'html.parser')
    
    paragraphs = []
    
    # Also capture headers
    for elem in soup.find_all(['h1', 'h2', 'h3', 'h4', 'p']):
        text = clean_text(elem.get_text())
        if not text:
            continue
        
        is_header = elem.name in ['h1', 'h2', 'h3', 'h4']
        is_hebrew = has_hebrew(text)
        
        paragraphs.append({
            'text': f"{{{{header:{text}}}}}" if is_header else text,
            'is_hebrew': is_hebrew,
            'is_header': is_header
        })
    
    return paragraphs


def convert_deuteronomy():
    """Convert Deuteronomy Keter Torah (Hebrew + English aligned by verse)."""
    print("\n" + "="*70)
    print("Converting: Deuteronomy Keter Torah Aaron ben Elijah")
    print("="*70)
    
    heb_file = HTML_DIR / "Deuteronomy_Keter_Torah_Aaron_ben_Elijah/Deuteronomy_Keter Torah_Aaron ben Elijah-Hebrew.html"
    eng_file = HTML_DIR / "Deuteronomy_Keter_Torah_Aaron_ben_Elijah/Deuteronomy_Keter Torah_Aaron ben Elijah-English.html"
    
    if not heb_file.exists() or not eng_file.exists():
        print("  ERROR: Files not found")
        return
    
    print("  Parsing Hebrew...")
    hebrew_verses = parse_verse_marked_html(heb_file)
    print(f"    Found {len(hebrew_verses)} verses/sections")
    
    print("  Parsing English...")
    english_verses = parse_verse_marked_html(eng_file)
    print(f"    Found {len(english_verses)} verses/sections")
    
    # Combine all verse markers
    all_verses = sorted(
        set(hebrew_verses.keys()) | set(english_verses.keys()),
        key=lambda x: (
            tuple(map(lambda p: float(p.split('-')[0]) if p.replace('-','').replace('.','').isdigit() else 999,
                     x.replace('_intro_', '999:').split(':')))
        )
    )
    
    print(f"  Total unique markers: {len(all_verses)}")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Text"
    ws.append(COLUMNS)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    line_num = 0
    for verse in all_verses:
        line_num += 1
        
        heb_text = hebrew_verses.get(verse, '')
        eng_text = english_verses.get(verse, '')
        
        # Add verse marker as comment if it's a real verse
        comment = verse if not verse.startswith('_') else ''
        
        ws.append([
            "Deuteronomy Keter Torah",  # File Name
            "",                          # Pattern
            "דברים כתר תורה",           # Hebrew Name
            "Deuteronomy Keter Torah",  # English Name
            "",                          # In Place of
            "",                          # Reciter
            line_num,                    # Hebrew Line #
            heb_text,                    # Hebrew Text
            "",                          # English Transliteration
            eng_text,                    # English Translation
            comment,                     # Comments (verse marker)
            "",                          # Time Starting
            ""                           # Time Ending
        ])
    
    # Adjust column widths
    ws.column_dimensions['H'].width = 80
    ws.column_dimensions['J'].width = 80
    ws.column_dimensions['K'].width = 15
    
    # Add Metadata sheet
    ws_meta = wb.create_sheet("Metadata")
    ws_meta.append(["Field", "Value"])
    ws_meta['A1'].font = Font(bold=True)
    ws_meta['B1'].font = Font(bold=True)
    
    ws_meta.append(["Title (English)", "Deuteronomy - Keter Torah"])
    ws_meta.append(["Title (Hebrew)", "דברים - כתר תורה"])
    ws_meta.append(["Author", "Aaron ben Elijah"])
    ws_meta.append(["Category", "Comments"])
    ws_meta.append(["Type", "Biblical Commentary"])
    ws_meta.append(["Total Verses", str(line_num)])
    ws_meta.append(["", ""])
    ws_meta.append(["--- FORMATTING ---", ""])
    ws_meta.append(["_text_", "Italics"])
    ws_meta.append(["**text**", "Bold"])
    ws_meta.append(["{{header:text}}", "Section Header"])
    
    ws_meta.column_dimensions['A'].width = 20
    ws_meta.column_dimensions['B'].width = 50
    
    # Save
    output_dir = OUTPUT_DIR / "Comments"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "Deuteronomy Keter Torah Aaron ben Elijah.xlsx"
    wb.save(output_path)
    print(f"  Created: {output_path}")


def convert_adderet_eliyahu():
    """Convert Adderet Eliyahu files (prose format)."""
    print("\n" + "="*70)
    print("Converting: Adderet Eliyahu (original chapters)")
    print("="*70)
    
    files = [
        ("Halakha_Adderet_Eliyahu_R_Elijah_Bashyatchi-0.html", "Introduction"),
        ("Halakha_Adderet_Eliyahu_R_Elijah_Bashyatchi-1.html", "Part 1"),
        ("Halakha_Adderet_Eliyahu_R_Elijah_Bashyatchi-2.html", "Part 2"),
    ]
    
    base_path = HTML_DIR / "Halakhah/Adderet_Eliyahu_R_Elijah_Bashyatchi/original"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Text"
    ws.append(COLUMNS)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    line_num = 0
    
    for filename, section in files:
        filepath = base_path / filename
        if not filepath.exists():
            print(f"  WARNING: {filename} not found")
            continue
        
        print(f"  Processing {section}...")
        paragraphs = parse_prose_html(filepath)
        print(f"    Found {len(paragraphs)} paragraphs")
        
        for para in paragraphs:
            line_num += 1
            text = para['text']
            
            if para['is_hebrew']:
                ws.append([
                    "Adderet Eliyahu",  # File Name
                    section,            # Pattern (section marker)
                    "אדרת אליהו",       # Hebrew Name
                    "Adderet Eliyahu",  # English Name
                    "",                 # In Place of
                    "",                 # Reciter
                    line_num,           # Hebrew Line #
                    text,               # Hebrew Text
                    "",                 # English Transliteration
                    "",                 # English Translation
                    "",                 # Comments
                    "",                 # Time Starting
                    ""                  # Time Ending
                ])
            else:
                ws.append([
                    "Adderet Eliyahu",
                    section,
                    "אדרת אליהו",
                    "Adderet Eliyahu",
                    "",
                    "",
                    line_num,
                    "",                 # Hebrew Text (empty)
                    "",                 # English Transliteration
                    text,               # English Translation
                    "",
                    "",
                    ""
                ])
    
    # Adjust column widths
    ws.column_dimensions['H'].width = 80
    ws.column_dimensions['J'].width = 80
    
    # Add Metadata sheet
    ws_meta = wb.create_sheet("Metadata")
    ws_meta.append(["Field", "Value"])
    ws_meta['A1'].font = Font(bold=True)
    ws_meta['B1'].font = Font(bold=True)
    
    ws_meta.append(["Title (English)", "Adderet Eliyahu"])
    ws_meta.append(["Title (Hebrew)", "אדרת אליהו"])
    ws_meta.append(["Author", "R. Elijah Bashyatchi"])
    ws_meta.append(["Category", "Halakhah"])
    ws_meta.append(["Type", "Legal Code"])
    ws_meta.append(["Total Paragraphs", str(line_num)])
    ws_meta.append(["Sections", "Introduction, Part 1, Part 2"])
    
    ws_meta.column_dimensions['A'].width = 20
    ws_meta.column_dimensions['B'].width = 50
    
    # Save
    output_dir = OUTPUT_DIR / "Halakhah"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "Adderet Eliyahu (Bashyatchi).xlsx"
    wb.save(output_path)
    print(f"  Created: {output_path}")


def convert_appendices():
    """Convert Adderet Eliyahu Appendices."""
    print("\n" + "="*70)
    print("Converting: Adderet Eliyahu (Combined Appendices)")
    print("="*70)
    
    filepath = HTML_DIR / "Halakhah/Adderet_Eliyahu_R_Elijah_Bashyatchi/Adderet Eliyahu (COMBINED APPENDICES) (SITE).html"
    
    if not filepath.exists():
        print("  ERROR: File not found")
        return
    
    print("  Parsing...")
    paragraphs = parse_prose_html(filepath)
    print(f"  Found {len(paragraphs)} paragraphs")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Text"
    ws.append(COLUMNS)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    line_num = 0
    for para in paragraphs:
        line_num += 1
        text = para['text']
        
        if para['is_hebrew']:
            ws.append([
                "Adderet Eliyahu Appendices",
                "",
                "אדרת אליהו - נספחים",
                "Adderet Eliyahu Appendices",
                "",
                "",
                line_num,
                text,
                "",
                "",
                "",
                "",
                ""
            ])
        else:
            ws.append([
                "Adderet Eliyahu Appendices",
                "",
                "אדרת אליהו - נספחים",
                "Adderet Eliyahu Appendices",
                "",
                "",
                line_num,
                "",
                "",
                text,
                "",
                "",
                ""
            ])
    
    ws.column_dimensions['H'].width = 80
    ws.column_dimensions['J'].width = 80
    
    # Metadata
    ws_meta = wb.create_sheet("Metadata")
    ws_meta.append(["Field", "Value"])
    ws_meta['A1'].font = Font(bold=True)
    ws_meta['B1'].font = Font(bold=True)
    
    ws_meta.append(["Title (English)", "Adderet Eliyahu - Appendices"])
    ws_meta.append(["Title (Hebrew)", "אדרת אליהו - נספחים"])
    ws_meta.append(["Category", "Halakhah"])
    ws_meta.append(["Total Paragraphs", str(line_num)])
    
    ws_meta.column_dimensions['A'].width = 20
    ws_meta.column_dimensions['B'].width = 50
    
    output_dir = OUTPUT_DIR / "Halakhah"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "Adderet Eliyahu (Appendices).xlsx"
    wb.save(output_path)
    print(f"  Created: {output_path}")


def main():
    print("PROSE HTML TO XLSX CONVERTER")
    print("="*70)
    
    convert_deuteronomy()
    convert_adderet_eliyahu()
    convert_appendices()
    
    print("\n" + "="*70)
    print("DONE!")
    print("="*70)


if __name__ == "__main__":
    main()
