#!/usr/bin/env python3
"""
Process XLSX files (source of truth) into JSON for the static site.
Only converts HTML when no XLSX exists for that text.
"""

import json
import re
from pathlib import Path
from openpyxl import load_workbook

# Source paths - XLSX is source of truth
XLSX_DIRS = [
    Path("/Users/shawn/knowledge-chatbot/documents/new_karaite/data_karaites/Out_xls"),
    Path("/Users/shawn/knowledge-chatbot/documents/new_karaite/data_karaites/HTML"),  # Some XLSX are here
    Path("/Users/shawn/knowledge-chatbot/documents/new_karaite/data_karaites/Word Documents"),
]
OUTPUT_DIR = Path("/Users/shawn/karaite-texts/data/texts")

def slugify(text):
    """Convert text to URL-friendly slug."""
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_-]+', '-', text)
    return text

def clean_text(text):
    """Clean whitespace from text."""
    if not text:
        return ""
    return re.sub(r'\s+', ' ', str(text)).strip()

def has_hebrew(text):
    """Check if text contains Hebrew characters."""
    return bool(re.search(r'[\u0590-\u05FF]', str(text))) if text else False

def get_category_from_path(filepath):
    """Extract category and subcategory from file path."""
    parts = filepath.parts
    
    # Find index of key folders
    for i, part in enumerate(parts):
        if part in ['Liturgy', 'Halakhah', 'Polemics', 'Comments', 'Exhortatory Literature']:
            category = part
            # Subcategory is next folder if exists
            if i + 1 < len(parts) - 1:  # -1 to exclude filename
                subcategory = parts[i + 1]
            else:
                subcategory = ""
            return category, subcategory
    
    # Default
    return "Liturgy", ""

def process_xlsx(filepath):
    """Process a single XLSX file into JSON format."""
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"    Error loading: {e}")
        return None
    
    results = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Get headers from first row
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[clean_text(cell.value).lower()] = col_idx
        
        # Find key columns
        col_hebrew = headers.get('hebrew text', headers.get('hebrew', 0))
        col_translit = headers.get('english transliteration', headers.get('transliteration', 0))
        col_english = headers.get('english translation', headers.get('english', 0))
        col_comments = headers.get('comments', 0)
        col_hebrew_name = headers.get('hebrew name', 0)
        col_english_name = headers.get('english name', 0)
        col_line_num = headers.get('hebrew line #', headers.get('line #', 0))
        
        if not col_hebrew and not col_translit:
            continue  # Skip sheets without text columns
        
        # Extract content
        content = []
        title_he = ""
        title_en = ""
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Get title from first row with data
            if col_hebrew_name and not title_he:
                val = row[col_hebrew_name - 1] if col_hebrew_name <= len(row) else None
                if val and has_hebrew(val):
                    title_he = clean_text(val)
            if col_english_name and not title_en:
                val = row[col_english_name - 1] if col_english_name <= len(row) else None
                if val and not has_hebrew(val):
                    title_en = clean_text(val)
            
            # Get text content
            hebrew = clean_text(row[col_hebrew - 1]) if col_hebrew and col_hebrew <= len(row) else ""
            translit = clean_text(row[col_translit - 1]) if col_translit and col_translit <= len(row) else ""
            english = clean_text(row[col_english - 1]) if col_english and col_english <= len(row) else ""
            comments = clean_text(row[col_comments - 1]) if col_comments and col_comments <= len(row) else ""
            
            if hebrew or translit:
                entry = {
                    "hebrew": hebrew,
                    "transliteration": translit,
                    "english": english,
                }
                if comments:
                    entry["comments"] = comments
                content.append(entry)
        
        if not content:
            continue
        
        # Use sheet name for title if not found in data
        if not title_en:
            title_en = sheet_name if sheet_name != 'Sheet1' else filepath.stem
        
        # Determine category from path
        category, subcategory = get_category_from_path(filepath)
        
        # Create unique ID
        base_id = slugify(title_en or sheet_name)
        if len(wb.sheetnames) > 1 and sheet_name != 'Sheet1':
            text_id = f"{slugify(filepath.stem)}-{slugify(sheet_name)}"
        else:
            text_id = base_id
        
        result = {
            "id": text_id,
            "title_en": title_en,
            "title_he": title_he,
            "category": category,
            "subcategory": subcategory,
            "content": content,
            "source_file": filepath.name,
            "source_sheet": sheet_name if len(wb.sheetnames) > 1 else None
        }
        
        results.append(result)
    
    return results

def build_catalog(texts):
    """Build the catalog.json index file."""
    catalog = {
        "categories": {},
        "texts": []
    }
    
    for text in texts:
        # Add to texts list
        catalog["texts"].append({
            "id": text["id"],
            "title_en": text["title_en"],
            "title_he": text["title_he"],
            "category": text["category"],
            "subcategory": text.get("subcategory", ""),
        })
        
        # Organize by category
        cat = text["category"]
        subcat = text.get("subcategory", "")
        
        if cat not in catalog["categories"]:
            catalog["categories"][cat] = {"subcategories": {}, "texts": []}
        
        if subcat and subcat != cat:
            if subcat not in catalog["categories"][cat]["subcategories"]:
                catalog["categories"][cat]["subcategories"][subcat] = []
            catalog["categories"][cat]["subcategories"][subcat].append(text["id"])
        else:
            catalog["categories"][cat]["texts"].append(text["id"])
    
    return catalog

def main():
    """Main processing function."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # Find all XLSX files
    all_xlsx = []
    for xlsx_dir in XLSX_DIRS:
        if xlsx_dir.exists():
            for xlsx_file in xlsx_dir.rglob("*.xlsx"):
                # Skip scratch, old, and temp files
                if 'scratch' in str(xlsx_file).lower():
                    continue
                if '_old' in xlsx_file.stem.lower():
                    continue
                if xlsx_file.stem.startswith('~'):
                    continue
                all_xlsx.append(xlsx_file)
    
    print(f"Found {len(all_xlsx)} XLSX files\n")
    
    processed_texts = []
    
    for xlsx_file in sorted(all_xlsx):
        print(f"  {xlsx_file.name}...", end=" ")
        
        results = process_xlsx(xlsx_file)
        if results:
            for text_data in results:
                # Save individual text file
                output_file = OUTPUT_DIR / f"{text_data['id']}.json"
                with open(output_file, 'w', encoding='utf-8') as f:
                    json.dump(text_data, f, ensure_ascii=False, indent=2)
                processed_texts.append(text_data)
            print(f"OK ({len(results)} text{'s' if len(results) > 1 else ''})")
        else:
            print("(no content)")
    
    # Build and save catalog
    print("\nBuilding catalog...")
    catalog = build_catalog(processed_texts)
    with open(OUTPUT_DIR.parent / "catalog.json", 'w', encoding='utf-8') as f:
        json.dump(catalog, f, ensure_ascii=False, indent=2)
    
    print(f"\nDone! Processed {len(processed_texts)} texts from {len(all_xlsx)} files.")
    print(f"Output: {OUTPUT_DIR}")

if __name__ == "__main__":
    main()
